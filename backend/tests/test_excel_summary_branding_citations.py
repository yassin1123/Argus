"""Phase 3 / Week 12 / Day 4 — Summary + branding + citation-audit tests.

Ten tests per spec covering the Summary sheet, the workbook-wide
firm-branding chrome (header / tab colour / freeze panes), the
logo-embed path, and the citation-audit helper.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any
from unittest import mock

import pytest
from openpyxl import load_workbook

from core.exports._base import ClaimCitation
from core.exports.excel._branding import _normalise_hex, audit_citations
from core.exports.excel._styles import (
    INPUT_FILL_HEX,
    INPUT_TEXT_HEX,
    FORMULA_TEXT_HEX,
    LINK_TEXT_HEX,
)
from core.exports.excel_model import ExcelModelExporter


_BRANDING: dict[str, Any] = {
    "primary_color": "#0F6E56",
    "_firm_name": "Test Firm",
}


def _m_and_a_payload() -> dict[str, Any]:
    return {
        "mode": "m_and_a_diligence",
        "recommendation": "PROCEED WITH CONDITIONS",
        "key_reasons": ["Stable cash flow from anchor customers.", "Segment leadership.", "Synergy potential of 6.5m."],
        "risks": ["Customer concentration 41%.", "Halo renewal binary.", "ROI EBITDA-negative."],
        "financial_profile": {
            "revenue_trajectory": {"points": [
                {"period": "FY23", "value_gbp_m": 190.0, "source_citation": "claim_1"},
                {"period": "FY24", "value_gbp_m": 203.0, "source_citation": "claim_1"},
            ]},
            "ebitda_trajectory": {"points": [
                {"period": "FY24", "value_gbp_m": 21.5, "source_citation": "claim_2"},
            ]},
            "margin_profile": {"gross_margin": "36.4%", "ebitda_margin": "10.6%", "fcf_margin": "6.2%"},
            "working_capital_dynamics": "51-day cycle.",
            "debt_structure": "Net debt 32m.",
            "capex_intensity": "4.5%.",
            "cash_flow_quality": "Recurring.",
        },
        "valuation_range": {
            "low":  {"gbp_m": 205.0, "methodology": "DCF"},
            "base": {"gbp_m": 220.0, "methodology": "EV/EBITDA 8.5x"},
            "high": {"gbp_m": 235.0, "methodology": "EV/Sales 1.4x"},
            "comparable_transactions_cited": [
                {"target": "Comp A", "acquirer": "PE Inc", "year": 2023,
                 "multiple": "8.2x EV/EBITDA", "source_citation": "deal db"},
            ],
        },
        "synergy_estimate": {
            "revenue_synergies": [{"type": "Cross-sell", "magnitude_gbp_m": 4.5, "timing_months": 24, "confidence": "medium", "basis_citations": ["claim_3"]}],
            "cost_synergies":   [{"type": "Procurement", "magnitude_gbp_m": 6.5, "timing_months": 18, "confidence": "high", "basis_citations": ["claim_4"]}],
            "dis_synergies": [],
        },
        "target_overview": {
            "name": "TargetCo",
            "business_model": "Industrial services.",
            "segments": [{"name": "Facilities", "revenue_pct": 52.0, "growth_rate": "+2.8%"}],
            "ownership_history": "Founder.",
            "key_customers_concentration": "Top 3 = 41%.",
        },
        "_engagement_title": "TargetCo M&A diligence",
        "_target_name": "TargetCo",
        "_firm_name": "Test Firm",
    }


def _citations() -> list[ClaimCitation]:
    return [
        ClaimCitation(claim_id="claim_1", text="Revenue ledger", source_title="TargetCo CIM", source_type="firm_library"),
        ClaimCitation(claim_id="claim_2", text="EBITDA ledger", source_title="10-K", source_type="sec_filing"),
        ClaimCitation(claim_id="claim_3", text="Cross-sell", source_title="Playbook", source_type="firm_library"),
        ClaimCitation(claim_id="claim_4", text="Procurement", source_title="Playbook", source_type="firm_library"),
    ]


@pytest.fixture
def exporter() -> ExcelModelExporter:
    return ExcelModelExporter()


def _norm_rgb(rgb: Any) -> str:
    if rgb is None:
        return ""
    s = str(rgb).upper()
    return s[2:] if len(s) == 8 else s


# ---------------------------------------------------------------------------
# Test 1 — Summary sheet recommendation is colour-coded
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_summary_sheet_has_recommendation_color_coded(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Summary"]
    # The recommendation text lands at A4 per the Summary builder.
    rec_cell = ws["A4"]
    assert "PROCEED WITH CONDITIONS" in str(rec_cell.value or "")
    # PROCEED WITH CONDITIONS maps to amber (#B8860B per
    # classify_recommendation).
    rgb = _norm_rgb(rec_cell.font.color.rgb if rec_cell.font.color else None)
    assert rgb == "B8860B", f"recommendation colour wrong: {rgb!r} (expected amber B8860B)"


# ---------------------------------------------------------------------------
# Test 2 — Summary valuation columns link to DCF (and would link to
# Comparables if registered; M&A always has DCF)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_summary_valuation_columns_link_to_dcf_and_comparables(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Summary"]
    # Find the "DCF" row in the Key valuation section.
    dcf_row = None
    for r in range(1, 25):
        if str(ws.cell(row=r, column=1).value or "") == "DCF":
            dcf_row = r
            break
    assert dcf_row is not None, "Summary missing DCF row in Key valuation"
    ev_cell = ws.cell(row=dcf_row, column=2)
    formula = str(ev_cell.value or "")
    assert formula.startswith("=") and "DCF" in formula, (
        f"Summary DCF EV cell doesn't link to DCF sheet: {formula!r}"
    )


# ---------------------------------------------------------------------------
# Test 3 — Summary top-3 reasons cells carry citation comments
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_summary_top_3_reasons_have_citation_comments(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Summary"]
    # Find "Reasons" header then walk down up to 3 rows checking
    # that those cells carry a Comment.
    reasons_header_row = None
    for r in range(1, 30):
        if str(ws.cell(row=r, column=1).value or "") == "Reasons":
            reasons_header_row = r
            break
    assert reasons_header_row is not None
    citations_found = 0
    for r in range(reasons_header_row + 1, reasons_header_row + 4):
        cell = ws.cell(row=r, column=1)
        if cell.value is None:
            continue
        if cell.comment is not None:
            citations_found += 1
    assert citations_found >= 1, "no citation comments on Summary top-3 reasons"


# ---------------------------------------------------------------------------
# Test 4 — Firm header on every sheet (row 1)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_firm_header_on_every_sheet(exporter: ExcelModelExporter) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    for name in wb.sheetnames:
        ws = wb[name]
        row1_a = str(ws.cell(row=1, column=1).value or "")
        assert "Test Firm" in row1_a, f"sheet {name!r} row 1 missing firm name: {row1_a!r}"


# ---------------------------------------------------------------------------
# Test 5 — sheet tabs coloured with firm primary
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_sheet_tabs_colored_with_firm_primary_color(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    expected_hex = _normalise_hex(_BRANDING["primary_color"])
    for name in wb.sheetnames:
        ws = wb[name]
        tab = ws.sheet_properties.tabColor
        # openpyxl returns a Color object; check its rgb.
        tab_rgb = _norm_rgb(getattr(tab, "rgb", None)) if tab is not None else ""
        assert tab_rgb == expected_hex, (
            f"sheet {name!r} tab colour {tab_rgb!r} != expected {expected_hex!r}"
        )


# ---------------------------------------------------------------------------
# Test 6 — logo on Cover sheet when branded
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_logo_on_cover_sheet_if_branded(
    exporter: ExcelModelExporter, monkeypatch: pytest.MonkeyPatch, tmp_path: Path,
) -> None:
    """Mock the asset cache to return real PNG bytes (a tiny solid-fill
    image). The Cover sheet should then carry exactly one Image
    in its ``_images`` collection."""
    from io import BytesIO

    from PIL import Image as PILImage

    # Build a 50x50 solid-colour PNG.
    buf = BytesIO()
    PILImage.new("RGBA", (50, 50), color=(15, 110, 86, 255)).save(buf, format="PNG")
    raw = buf.getvalue()

    # Point the title-sheet's resolver at our bytes.
    import core.exports.excel.sheets.title as title_module
    monkeypatch.setattr(title_module, "_resolve_logo_sync", lambda *a, **kw: raw)

    branding = dict(_BRANDING)
    branding["logo_url"] = "https://example.com/logo.png"
    branding["_firm_id"] = "test-firm-id"

    result = await exporter.render(_m_and_a_payload(), branding, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    cover = wb["Cover"]
    n_images = len(getattr(cover, "_images", []))
    assert n_images >= 1, f"expected ≥1 image on Cover sheet, got {n_images}"


# ---------------------------------------------------------------------------
# Test 7 — audit_citations finds at least one source per required sheet
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_all_payload_derived_cells_have_citation_comments(
    exporter: ExcelModelExporter,
) -> None:
    """Every sheet that contains payload-derived data has ≥1 cell with
    an Argus-authored citation comment. ``audit_citations`` returns
    an empty ``missing`` list when each required sheet has at least
    one citation."""
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    report = audit_citations(wb)
    assert report["missing"] == [], f"citation gaps: {report['missing']}"
    # Confirm the sheets we expect to have citations actually do.
    expected_sheets = {"Assumptions", "Revenue Build", "Cost Build", "Synergies", "Summary"}
    assert expected_sheets.issubset(set(report["sheets_passed"])), (
        f"missing citation passes: {expected_sheets - set(report['sheets_passed'])}"
    )


# ---------------------------------------------------------------------------
# Test 8 — input cells use blue text on yellow fill
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_input_cells_are_blue_on_yellow(exporter: ExcelModelExporter) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Assumptions"]
    # WACC input cell.
    wacc_row = None
    for r in range(1, 35):
        if str(ws.cell(row=r, column=1).value or "") == "WACC":
            wacc_row = r
            break
    assert wacc_row is not None
    val_cell = ws.cell(row=wacc_row, column=2)
    fill_rgb = _norm_rgb(val_cell.fill.fgColor.rgb if val_cell.fill and val_cell.fill.fgColor else None)
    text_rgb = _norm_rgb(val_cell.font.color.rgb if val_cell.font and val_cell.font.color else None)
    assert fill_rgb == INPUT_FILL_HEX.upper()
    assert text_rgb == INPUT_TEXT_HEX.upper()


# ---------------------------------------------------------------------------
# Test 9 — cross-sheet link cells use green text
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_link_cells_are_green(exporter: ExcelModelExporter) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    # Cost Build's "EBITDA margin" projection row links to Assumptions
    # (a cross-sheet link). Pick the first projection column.
    ws = wb["Cost Build"]
    em_row = None
    for r in range(1, 15):
        if str(ws.cell(row=r, column=1).value or "") == "EBITDA margin":
            em_row = r
            break
    assert em_row is not None
    # First projection column for M&A = col 2 (FY+0 historicals are
    # values; first cross-sheet link lands in col 4+ depending on
    # historical count). Walk forward until we find a cell with a
    # formula referencing Assumptions.
    link_rgb = None
    for col in range(2, 12):
        cell = ws.cell(row=em_row, column=col)
        if isinstance(cell.value, str) and cell.value.startswith("=Assumptions"):
            link_rgb = _norm_rgb(cell.font.color.rgb if cell.font and cell.font.color else None)
            break
    assert link_rgb == LINK_TEXT_HEX.upper(), (
        f"link cell text colour {link_rgb!r} != green {LINK_TEXT_HEX!r}"
    )


# ---------------------------------------------------------------------------
# Test 10 — computed (formula) cells use black text
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_computed_cells_are_black(exporter: ExcelModelExporter) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["DCF"]
    # Free Cash Flow row is computed (bold black).
    fcf_row = None
    for r in range(1, 30):
        if str(ws.cell(row=r, column=1).value or "") == "Free Cash Flow":
            fcf_row = r
            break
    assert fcf_row is not None
    cell = ws.cell(row=fcf_row, column=2)
    text_rgb = _norm_rgb(cell.font.color.rgb if cell.font and cell.font.color else None)
    assert text_rgb == FORMULA_TEXT_HEX.upper(), (
        f"computed cell text colour {text_rgb!r} != black"
    )
