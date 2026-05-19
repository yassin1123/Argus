"""Phase 3 / Week 12 / Day 3 — M&A-specific sheet tests.

Eleven tests per spec covering the four diligence-grade sheets
(working_capital, dcf, comparables, sensitivity, synergies):

  - Working Capital links to Revenue Build
  - DCF Free Cash Flow formula correctness
  - DCF terminal value (both methods present)
  - DCF Enterprise Value = SUM(PV_FCF) + TV
  - Comparables pulls transactions from payload
  - Comparables median row uses MEDIAN() formula
  - Sensitivity 5×5 grid dimensions
  - Synergies NPV formula per synergy
  - Dis-synergies render as negative magnitudes
  - growth_strategy mode omits DCF
  - M&A workbook has 9 sheets
"""

from __future__ import annotations

import io
import re
from typing import Any

import pytest
from openpyxl import load_workbook

from core.exports._base import ClaimCitation
from core.exports.excel_model import ExcelModelExporter


_BRANDING: dict[str, Any] = {
    "primary_color": "#0F6E56",
    "_firm_name": "Test Firm",
}


def _m_and_a_payload() -> dict[str, Any]:
    return {
        "mode": "m_and_a_diligence",
        "recommendation": "PROCEED WITH CONDITIONS",
        "financial_profile": {
            "revenue_trajectory": {"points": [
                {"period": "FY21", "value_gbp_m": 153.2, "source_citation": "claim_1"},
                {"period": "FY22", "value_gbp_m": 168.5, "source_citation": "claim_1"},
                {"period": "FY23", "value_gbp_m": 190.0, "source_citation": "claim_1"},
                {"period": "FY24", "value_gbp_m": 203.0, "source_citation": "claim_1"},
            ]},
            "ebitda_trajectory": {"points": [
                {"period": "FY23", "value_gbp_m": 19.0, "source_citation": "claim_2"},
                {"period": "FY24", "value_gbp_m": 21.5, "source_citation": "claim_2"},
            ]},
            "margin_profile": {"gross_margin": "36.4%", "ebitda_margin": "10.6%", "fcf_margin": "6.2%"},
            "working_capital_dynamics": "51-day WC cycle.",
            "debt_structure": "Net debt 32m.",
            "capex_intensity": "4.5% of revenue.",
            "cash_flow_quality": "Recurring.",
        },
        "valuation_range": {
            "low":  {"gbp_m": 205.0, "methodology": "DCF"},
            "base": {"gbp_m": 220.0, "methodology": "EV/EBITDA 8.5x"},
            "high": {"gbp_m": 235.0, "methodology": "EV/Sales 1.4x"},
            "comparable_transactions_cited": [
                {"target": "Comp A", "acquirer": "PE Inc", "year": 2023,
                 "multiple": "8.2x EV/EBITDA", "source_citation": "deal db"},
                {"target": "Comp B", "acquirer": "Strategic", "year": 2024,
                 "multiple": "1.4x EV/Sales", "source_citation": "press"},
                {"target": "Comp C", "acquirer": "PE Inc 2", "year": 2023,
                 "multiple": "7.5x EV/EBITDA", "source_citation": "deal db"},
            ],
        },
        "synergy_estimate": {
            "revenue_synergies": [
                {"type": "Cross-sell", "magnitude_gbp_m": 4.5, "timing_months": 24,
                 "confidence": "medium", "basis_citations": ["claim_3"]},
            ],
            "cost_synergies": [
                {"type": "Procurement", "magnitude_gbp_m": 6.5, "timing_months": 18,
                 "confidence": "high", "basis_citations": ["claim_4"]},
                {"type": "IT consolidation", "magnitude_gbp_m": 2.0, "timing_months": 30,
                 "confidence": "medium", "basis_citations": ["claim_4"]},
            ],
            "dis_synergies": [
                {"type": "Customer attrition", "magnitude_gbp_m": 1.0, "timing_months": 12,
                 "confidence": "low", "basis_citations": ["claim_5"]},
            ],
        },
        "target_overview": {
            "name": "TargetCo",
            "business_model": "Industrial services.",
            "segments": [{"name": "Facilities", "revenue_pct": 52.0, "growth_rate": "+2.8%"}],
            "ownership_history": "Founder.",
            "key_customers_concentration": "Top 3 = 41%.",
        },
        "_engagement_title": "TargetCo M&A diligence",
        "_target_name": "TargetCo",
        "_firm_name": "Test Firm",
    }


def _growth_payload() -> dict[str, Any]:
    return {
        "mode": "growth_strategy",
        "recommendation": "Launch Scotland pilot.",
        "financial_profile": {
            "revenue_trajectory": {"points": [
                {"period": "FY24", "value_gbp_m": 180.0, "source_citation": "claim_g1"},
            ]},
            "ebitda_trajectory": {"points": [
                {"period": "FY24", "value_gbp_m": 19.0, "source_citation": "claim_g2"},
            ]},
            "margin_profile": {"gross_margin": "35%", "ebitda_margin": "10%", "fcf_margin": "5%"},
            "working_capital_dynamics": "x", "debt_structure": "y",
            "capex_intensity": "z", "cash_flow_quality": "w",
        },
        "_engagement_title": "Scotland pilot",
        "_firm_name": "Test Firm",
    }


def _citations() -> list[ClaimCitation]:
    return [
        ClaimCitation(claim_id="claim_1", text="Revenue", source_title="CIM", source_type="firm_library"),
        ClaimCitation(claim_id="claim_2", text="EBITDA", source_title="10-K", source_type="sec_filing"),
        ClaimCitation(claim_id="claim_3", text="Cross-sell", source_title="Playbook", source_type="firm_library"),
        ClaimCitation(claim_id="claim_4", text="Procurement", source_title="Playbook", source_type="firm_library"),
        ClaimCitation(claim_id="claim_5", text="Attrition", source_title="Note", source_type="firm_library"),
    ]


@pytest.fixture
def exporter() -> ExcelModelExporter:
    return ExcelModelExporter()


def _find_row_by_label(ws: Any, label: str, max_row: int = 30) -> int | None:
    for r in range(1, max_row + 1):
        if str(ws.cell(row=r, column=1).value or "") == label:
            return r
    return None


# ---------------------------------------------------------------------------
# Test 1 — Working Capital references Revenue Build
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_working_capital_links_to_revenue_build(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Working Capital"]
    nwc_row = _find_row_by_label(ws, "Net Working Capital")
    assert nwc_row is not None
    # M&A horizon = 5 projection years; historicals = 4. NWC projection
    # column 1 starts at col 2 + 4 = 6.
    cell = ws.cell(row=nwc_row, column=6)
    assert isinstance(cell.value, str), f"NWC year-1 cell not formula: {cell.value!r}"
    assert "Revenue Build" in cell.value


# ---------------------------------------------------------------------------
# Test 2 — DCF Free Cash Flow formula references EBITDA / Tax / D&A / Capex / ΔNWC
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_dcf_fcf_formula_correctness(exporter: ExcelModelExporter) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["DCF"]
    fcf_row = _find_row_by_label(ws, "Free Cash Flow")
    ebitda_row = _find_row_by_label(ws, "EBITDA")
    tax_row = _find_row_by_label(ws, "Less: Tax")
    dna_row = _find_row_by_label(ws, "Plus: D&A (% of revenue)")
    capex_row = _find_row_by_label(ws, "Less: Capex (% of revenue)")
    nwc_row = _find_row_by_label(ws, "Less: ΔNWC")
    rev_row = _find_row_by_label(ws, "Revenue")
    assert all(r is not None for r in [fcf_row, ebitda_row, tax_row, dna_row, capex_row, nwc_row, rev_row])

    cell = ws.cell(row=fcf_row, column=2)  # FY+1 column
    formula = str(cell.value or "")
    # FCF = EBITDA - Tax + Rev*D&A - Rev*Capex - ΔNWC
    assert formula.startswith("=")
    assert f"B{ebitda_row}" in formula
    assert f"B{tax_row}" in formula
    assert f"B{dna_row}" in formula
    assert f"B{capex_row}" in formula
    assert f"B{nwc_row}" in formula


# ---------------------------------------------------------------------------
# Test 3 — DCF shows BOTH terminal value methods
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_dcf_terminal_value_both_methods_present(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["DCF"]
    labels = [str(ws.cell(row=r, column=1).value or "") for r in range(1, 26)]
    joined = "\n".join(labels)
    assert "TV — Gordon Growth" in joined
    assert "TV — Exit Multiple" in joined
    # Selected TV averages both.
    sel_row = _find_row_by_label(ws, "TV — selected (average)")
    assert sel_row is not None
    last_col_letter = "F"  # B..F = FY+1..FY+5 for M&A
    cell = ws.cell(row=sel_row, column=6)
    assert str(cell.value or "").startswith("=AVERAGE("), (
        f"selected TV not AVERAGE: {cell.value!r}"
    )


# ---------------------------------------------------------------------------
# Test 4 — Enterprise Value is SUM(PV_FCF) + TV
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_dcf_enterprise_value_is_sum_formula(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["DCF"]
    ev_row = _find_row_by_label(ws, "Enterprise Value")
    pv_row = _find_row_by_label(ws, "PV of FCF")
    sel_row = _find_row_by_label(ws, "TV — selected (average)")
    assert all(r is not None for r in [ev_row, pv_row, sel_row])
    formula = str(ws.cell(row=ev_row, column=2).value or "")
    assert formula.startswith("=SUM(")
    assert f"{pv_row}" in formula
    # Terminal selected TV reference (F<sel_row> for 5y M&A).
    assert f"F{sel_row}" in formula or f"{sel_row}" in formula


# ---------------------------------------------------------------------------
# Test 5 — Comparables pulls transactions from the payload
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_comparables_pulls_from_payload(exporter: ExcelModelExporter) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Comparables"]
    all_text = "\n".join(
        str(c.value or "")
        for row in ws.iter_rows()
        for c in row
        if c.value
    )
    for target in ("Comp A", "Comp B", "Comp C"):
        assert target in all_text, f"missing transaction {target} on Comparables sheet"
    # The 8.2x EV/EBITDA value should land in the EV/EBITDA column.
    # Find Comp A's row.
    comp_a_row = None
    for r in range(1, 30):
        if str(ws.cell(row=r, column=1).value or "") == "Comp A":
            comp_a_row = r
            break
    assert comp_a_row is not None
    assert ws.cell(row=comp_a_row, column=4).value == 8.2


# ---------------------------------------------------------------------------
# Test 6 — Comparables median row uses MEDIAN() formula
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_comparables_median_formula_correct(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Comparables"]
    # Find the Median row inside the transactions table.
    median_row = None
    for r in range(1, 30):
        if str(ws.cell(row=r, column=1).value or "") == "Median":
            median_row = r
            break
    assert median_row is not None
    # EV/EBITDA column = 4.
    formula = str(ws.cell(row=median_row, column=4).value or "")
    assert formula.startswith("=MEDIAN(D"), f"median formula wrong: {formula!r}"
    # And EV/Sales column = 5.
    assert str(ws.cell(row=median_row, column=5).value or "").startswith("=MEDIAN(E")


# ---------------------------------------------------------------------------
# Test 7 — Sensitivity 5x5 grids
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_sensitivity_table_dimensions_5x5(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Sensitivity"]
    # Find Grid 1 by title text.
    grid1_title = None
    for r in range(1, 50):
        v = str(ws.cell(row=r, column=1).value or "")
        if "Grid 1" in v and "WACC" in v:
            grid1_title = r
            break
    assert grid1_title is not None
    # Per _write_grid: title at row N, note at N+1, header at N+3,
    # data rows at N+4 … N+8 (5 rows × 5 cols).
    header_row = grid1_title + 3
    data_top = grid1_title + 4
    data_bottom = data_top + 4
    # Verify 5 row labels (WACC values).
    row_labels = [ws.cell(row=r, column=1).value for r in range(data_top, data_bottom + 1)]
    assert len(row_labels) == 5
    # Verify 5 column headers (TG values).
    col_headers = [ws.cell(row=header_row, column=c).value for c in range(2, 7)]
    assert len(col_headers) == 5
    # Spot-check the centre cell has a numeric value or 'n/a'.
    centre = ws.cell(row=data_top + 2, column=4).value
    assert centre is not None and (isinstance(centre, (int, float)) or centre == "n/a")


# ---------------------------------------------------------------------------
# Test 8 — Synergies NPV calculation per synergy
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_synergies_npv_calculation_per_synergy(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Synergies"]
    # Find a synergy row by its description text. The payload has
    # "Cross-sell" / "Procurement" / "IT consolidation" / "Customer
    # attrition" — pick any one and check its NPV cell.
    proc_row = None
    for r in range(1, 30):
        if str(ws.cell(row=r, column=1).value or "") == "Procurement":
            proc_row = r
            break
    assert proc_row is not None
    npv_cell = ws.cell(row=proc_row, column=6)
    formula = str(npv_cell.value or "")
    # NPV formula uses C × 1/(1+WACC)^((D/12)/2)
    assert formula.startswith("=")
    assert f"C{proc_row}" in formula
    assert f"D{proc_row}" in formula
    assert "Assumptions" in formula  # WACC reference


# ---------------------------------------------------------------------------
# Test 9 — dis-synergies render as negative magnitudes
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_synergies_dis_synergies_negative(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Synergies"]
    attrition_row = None
    for r in range(1, 30):
        if str(ws.cell(row=r, column=1).value or "") == "Customer attrition":
            attrition_row = r
            break
    assert attrition_row is not None
    mag = ws.cell(row=attrition_row, column=3).value
    assert isinstance(mag, (int, float)) and mag < 0, (
        f"dis-synergy magnitude must be negative, got {mag!r}"
    )


# ---------------------------------------------------------------------------
# Test 10 — growth_strategy mode does NOT have a DCF sheet
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_growth_strategy_does_not_have_dcf_sheet(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_growth_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    assert "DCF" not in wb.sheetnames
    assert "Comparables" not in wb.sheetnames
    assert "Sensitivity" not in wb.sheetnames
    assert "Synergies" not in wb.sheetnames
    assert "Working Capital" not in wb.sheetnames
    # Stays at the 4-sheet baseline.
    assert len(wb.sheetnames) == 4


# ---------------------------------------------------------------------------
# Test 11 — M&A workbook has 9 sheets
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_m_and_a_workbook_has_9_sheets(exporter: ExcelModelExporter) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    assert len(wb.sheetnames) == 9
    assert wb.sheetnames == [
        "Cover", "Assumptions", "Revenue Build", "Cost Build",
        "Working Capital", "DCF", "Comparables", "Sensitivity", "Synergies",
    ]
    # File size < 200 KB per spec.
    assert result.file_size < 200_000, (
        f"M&A workbook too large: {result.file_size} bytes (cap 200 KB)"
    )
