"""Phase 3 / Week 12 / Day 2 — Revenue Build + Cost Build tests.

Eight tests per spec covering:
  - historical-from-payload + formula-not-value projection cells
  - M&A segment-level detail + growth single-line shape
  - fallback when financial_profile.revenue_trajectory is missing
  - cross-sheet references (Cost Build → Revenue Build)
  - EBITDA formula correctness (string-level — openpyxl can't
    evaluate formulas itself)
  - number-format application
"""

from __future__ import annotations

import io
import re
from typing import Any

import pytest
from openpyxl import load_workbook

from core.exports._base import ClaimCitation
from core.exports.excel._refs import NUMBER_FORMATS
from core.exports.excel_model import ExcelModelExporter


_BRANDING: dict[str, Any] = {
    "primary_color": "#0F6E56",
    "_firm_name": "Test Firm",
}


def _m_and_a_payload() -> dict[str, Any]:
    return {
        "mode": "m_and_a_diligence",
        "recommendation": "PROCEED WITH CONDITIONS",
        "financial_profile": {
            "revenue_trajectory": {"points": [
                {"period": "FY21", "value_gbp_m": 153.2, "source_citation": "claim_1"},
                {"period": "FY22", "value_gbp_m": 168.5, "source_citation": "claim_1"},
                {"period": "FY23", "value_gbp_m": 190.0, "source_citation": "claim_1"},
                {"period": "FY24", "value_gbp_m": 203.0, "source_citation": "claim_1"},
            ]},
            "ebitda_trajectory": {"points": [
                {"period": "FY23", "value_gbp_m": 19.0, "source_citation": "claim_2"},
                {"period": "FY24", "value_gbp_m": 21.5, "source_citation": "claim_2"},
            ]},
            "margin_profile": {
                "gross_margin": "36.4%", "ebitda_margin": "10.6%", "fcf_margin": "6.2%",
            },
            "working_capital_dynamics": "x", "debt_structure": "y",
            "capex_intensity": "z", "cash_flow_quality": "w",
        },
        "valuation_range": {
            "low":  {"gbp_m": 205.0, "methodology": "DCF"},
            "base": {"gbp_m": 220.0, "methodology": "EV/EBITDA"},
            "high": {"gbp_m": 235.0, "methodology": "EV/Sales"},
        },
        "target_overview": {
            "name": "TargetCo",
            "business_model": "Industrial services.",
            "segments": [
                {"name": "Facilities maintenance", "revenue_pct": 52.0, "growth_rate": "+2.8%"},
                {"name": "Mechanical services",   "revenue_pct": 28.0, "growth_rate": "+1.2%"},
                {"name": "Project services",      "revenue_pct": 20.0, "growth_rate": "+0.5%"},
            ],
            "geographies": [{"geography": "UK", "revenue_pct": 91.0}],
            "ownership_history": "Founder-owned.",
            "key_customers_concentration": "Top 3 = 41%.",
        },
        "_engagement_title": "TargetCo M&A diligence",
        "_firm_name": "Test Firm",
    }


def _growth_payload() -> dict[str, Any]:
    return {
        "mode": "growth_strategy",
        "recommendation": "Launch Scotland pilot.",
        "financial_profile": {
            "revenue_trajectory": {"points": [
                {"period": "FY23", "value_gbp_m": 175.0, "source_citation": "claim_g1"},
                {"period": "FY24", "value_gbp_m": 180.0, "source_citation": "claim_g1"},
            ]},
            "ebitda_trajectory": {"points": [
                {"period": "FY24", "value_gbp_m": 19.0, "source_citation": "claim_g2"},
            ]},
            "margin_profile": {"gross_margin": "35%", "ebitda_margin": "10%", "fcf_margin": "5%"},
            "working_capital_dynamics": "x", "debt_structure": "y",
            "capex_intensity": "z", "cash_flow_quality": "w",
        },
        "_engagement_title": "TargetCo Scotland pilot",
        "_firm_name": "Test Firm",
    }


def _citations() -> list[ClaimCitation]:
    return [
        ClaimCitation(
            claim_id="claim_1", text="Revenue ledger",
            source_title="TargetCo CIM", source_type="firm_library",
        ),
        ClaimCitation(
            claim_id="claim_2", text="EBITDA ledger",
            source_title="10-K 2023", source_type="sec_filing",
        ),
        ClaimCitation(
            claim_id="claim_g1", text="Growth revenue",
            source_title="Growth pack", source_type="firm_library",
        ),
    ]


@pytest.fixture
def exporter() -> ExcelModelExporter:
    return ExcelModelExporter()


# ---------------------------------------------------------------------------
# Test 1 — historical cells pull values from the payload trajectory
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_revenue_build_historical_pulls_from_payload(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Revenue Build"]
    # Year header row (row 3) lists the 4 historical periods.
    header = [ws.cell(row=3, column=col).value for col in range(2, 6)]
    assert header == ["FY21", "FY22", "FY23", "FY24"]

    # Find the "Total revenue" row; its historical cells should sum
    # to roughly the payload's totals.
    total_row = None
    for r in range(4, 20):
        v = ws.cell(row=r, column=1).value
        if v and "Total revenue" in str(v):
            total_row = r
            break
    assert total_row is not None, "Total revenue row missing"
    # Historical totals should be formulas (SUM(...)) covering segments.
    for col in range(2, 6):
        cell = ws.cell(row=total_row, column=col)
        assert isinstance(cell.value, str) and cell.value.startswith("=SUM(")


# ---------------------------------------------------------------------------
# Test 2 — projection cells contain formulas, not literal numbers
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_revenue_build_projection_uses_formula_not_value(
    exporter: ExcelModelExporter,
) -> None:
    """Consulting modelling discipline: historicals are values +
    citations; projections are formulas."""
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Revenue Build"]
    # M&A → 4 historical + 5 projection = cols 2..10.
    # Find one segment row and inspect cols 6..10 (projection).
    seg_row = None
    for r in range(4, 12):
        label = str(ws.cell(row=r, column=1).value or "")
        if "Facilities maintenance" in label:
            seg_row = r
            break
    assert seg_row is not None
    for col in range(6, 11):
        cell = ws.cell(row=seg_row, column=col)
        assert isinstance(cell.value, str) and cell.value.startswith("=")
        # Formula references the previous cell × (1 + growth).
        assert re.search(r"\*\(1\+", cell.value), (
            f"projection cell {cell.coordinate} formula doesn't look like growth: {cell.value!r}"
        )


# ---------------------------------------------------------------------------
# Test 3 — M&A mode renders per-segment rows
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_revenue_build_segment_detail_for_m_and_a(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Revenue Build"]
    labels = [str(ws.cell(row=r, column=1).value or "") for r in range(4, 12)]
    joined = "\n".join(labels)
    assert "Facilities maintenance" in joined
    assert "Mechanical services" in joined
    assert "Project services" in joined
    assert "Total revenue" in joined


# ---------------------------------------------------------------------------
# Test 4 — growth_strategy renders a single revenue row, no segments
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_revenue_build_no_segments_for_growth(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_growth_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Revenue Build"]
    labels = [str(ws.cell(row=r, column=1).value or "") for r in range(4, 10)]
    joined = "\n".join(labels)
    assert "Revenue" in joined
    # M&A-only segment names must NOT appear.
    for seg in ("Facilities maintenance", "Mechanical services", "Project services"):
        assert seg not in joined, f"growth deck wrongly contains segment {seg!r}"
    assert "Total revenue" not in joined
    # Year header has 2 historical + 3 projection = cols 2..6.
    header = [ws.cell(row=3, column=col).value for col in range(2, 7)]
    assert header == ["FY23", "FY24", "FY+1", "FY+2", "FY+3"]


# ---------------------------------------------------------------------------
# Test 5 — fallback path when trajectory is missing
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_revenue_build_falls_back_when_trajectory_missing(
    exporter: ExcelModelExporter,
) -> None:
    """When payload.financial_profile.revenue_trajectory is empty,
    historicals render a placeholder string + projections become
    BLUE input cells defaulting to 0."""
    payload = {
        "mode": "growth_strategy",
        "recommendation": "Test fallback.",
        "_engagement_title": "Fallback engagement",
        "_firm_name": "Test Firm",
        # No financial_profile at all.
    }
    result = await exporter.render(payload, _BRANDING, [])
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Revenue Build"]
    placeholder_found = False
    n_yellow_inputs = 0

    def _norm_rgb(rgb: Any) -> str:
        if rgb is None:
            return ""
        s = str(rgb).upper()
        return s[2:] if len(s) == 8 else s

    for row in ws.iter_rows():
        for cell in row:
            v = str(cell.value or "")
            if "not produced for this engagement" in v.lower():
                placeholder_found = True
            try:
                fg = cell.fill.fgColor
                if fg is not None and fg.rgb and _norm_rgb(fg.rgb) == "FFFF99":
                    n_yellow_inputs += 1
            except Exception:
                pass
    assert placeholder_found, "fallback placeholder string missing on Revenue Build"
    # 3 projection years for growth_strategy → 3 yellow input cells.
    assert n_yellow_inputs >= 3, (
        f"expected ≥3 yellow input cells in fallback path, got {n_yellow_inputs}"
    )


# ---------------------------------------------------------------------------
# Test 6 — Cost Build references Revenue Build cells via formula
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_cost_build_references_revenue_build_via_formula(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Cost Build"]
    # Find the "Revenue" row (in the projection block).
    revenue_row = None
    for r in range(4, 15):
        label = str(ws.cell(row=r, column=1).value or "")
        if label == "Revenue":
            revenue_row = r
            break
    assert revenue_row is not None, "Cost Build missing Revenue row in projection block"
    # First projection column for M&A is col 2 + 4 historical = 6.
    cell = ws.cell(row=revenue_row, column=6)
    assert isinstance(cell.value, str), f"Revenue row col 6 not a formula: {cell.value!r}"
    assert "Revenue Build" in cell.value
    assert cell.value.startswith("=")


# ---------------------------------------------------------------------------
# Test 7 — EBITDA formula correctness (= Revenue × EBITDA margin)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_ebitda_margin_formula_correctness(
    exporter: ExcelModelExporter,
) -> None:
    """openpyxl doesn't evaluate formulas; assert the formula string
    pattern instead. EBITDA cell on the Cost Build projection block
    must equal ``=<col><rev_row>*<col><em_row>``."""
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Cost Build"]

    rows: dict[str, int] = {}
    for r in range(4, 15):
        label = str(ws.cell(row=r, column=1).value or "")
        if label in ("Revenue", "EBITDA margin", "EBITDA"):
            rows[label] = r
    assert "Revenue" in rows and "EBITDA margin" in rows and "EBITDA" in rows
    rev_r = rows["Revenue"]
    em_r = rows["EBITDA margin"]
    ebitda_r = rows["EBITDA"]

    # M&A first projection col = 6 (col B + 4 historicals = col F).
    # Use openpyxl's get_column_letter for portability.
    from openpyxl.utils import get_column_letter
    proj_col_letter = get_column_letter(6)
    ebitda_cell = ws.cell(row=ebitda_r, column=6)
    assert ebitda_cell.value == (
        f"={proj_col_letter}{rev_r}*{proj_col_letter}{em_r}"
    ), f"EBITDA formula wrong: {ebitda_cell.value!r}"


# ---------------------------------------------------------------------------
# Test 8 — number formats applied per cell type
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_number_format_applied_per_cell_type(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_m_and_a_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))

    # Revenue Build historical cells: currency_gbp_m.
    rb = wb["Revenue Build"]
    seg_row = None
    for r in range(4, 12):
        label = str(rb.cell(row=r, column=1).value or "")
        if "Facilities maintenance" in label:
            seg_row = r
            break
    assert seg_row is not None
    fmt = rb.cell(row=seg_row, column=2).number_format
    assert fmt == NUMBER_FORMATS["currency_gbp_m"], (
        f"historical revenue cell number_format = {fmt!r}, expected currency_gbp_m"
    )

    # Cost Build projection EBITDA margin cell: percent format.
    cb = wb["Cost Build"]
    em_row = None
    for r in range(4, 15):
        if str(cb.cell(row=r, column=1).value or "") == "EBITDA margin":
            em_row = r
            break
    assert em_row is not None
    fmt = cb.cell(row=em_row, column=6).number_format
    assert fmt == NUMBER_FORMATS["percent"], (
        f"EBITDA margin projection cell number_format = {fmt!r}, expected percent"
    )

    # Assumptions input cell number_format is percent.
    ass = wb["Assumptions"]
    wacc_row = None
    for r in range(1, 30):
        if str(ass.cell(row=r, column=1).value or "") == "WACC":
            wacc_row = r
            break
    assert wacc_row is not None
    fmt = ass.cell(row=wacc_row, column=2).number_format
    assert fmt == NUMBER_FORMATS["percent"], (
        f"WACC input cell number_format = {fmt!r}, expected percent"
    )
