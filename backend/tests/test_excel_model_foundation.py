"""Phase 3 / Week 12 / Day 1 — Excel exporter foundation tests.

Six tests per spec covering the XLSX round-trip, sheet content, the
industry-standard input-cell colour convention, default consultant
assumptions, citation comments on payload-derived cells, and on-disk
reopen.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from core.exports._base import ClaimCitation
from core.exports.excel._styles import INPUT_FILL_HEX, INPUT_TEXT_HEX
from core.exports.excel_model import ExcelModelExporter


_BRANDING: dict[str, Any] = {
    "primary_color": "#0F6E56",
    "_firm_name": "Test Firm",
}


def _payload() -> dict[str, Any]:
    return {
        "mode": "m_and_a_diligence",
        "recommendation": "PROCEED WITH CONDITIONS",
        "summary": "Stable.",
        "key_reasons": ["r1"],
        "risks": ["x1"],
        "sources": [{"type": "firm_library", "title": "Playbook"}],
        "financial_profile": {
            "revenue_trajectory": {"points": [
                {"period": "FY21", "value_gbp_m": 153.2, "source_citation": "claim_1"},
                {"period": "FY22", "value_gbp_m": 168.5, "source_citation": "claim_1"},
                {"period": "FY23", "value_gbp_m": 190.0, "source_citation": "claim_1"},
                {"period": "FY24", "value_gbp_m": 203.0, "source_citation": "claim_1"},
            ]},
            "ebitda_trajectory": {"points": [
                {"period": "FY23", "value_gbp_m": 19.0, "source_citation": "claim_2"},
                {"period": "FY24", "value_gbp_m": 21.5, "source_citation": "claim_2"},
            ]},
            "margin_profile": {"gross_margin": "36.4%", "ebitda_margin": "10.6%", "fcf_margin": "6.2%"},
            "working_capital_dynamics": "x", "debt_structure": "y",
            "capex_intensity": "z", "cash_flow_quality": "w",
        },
        "valuation_range": {
            "low":  {"gbp_m": 205.0, "methodology": "DCF @ WACC 10%"},
            "base": {"gbp_m": 220.0, "methodology": "EV/EBITDA 8.5x"},
            "high": {"gbp_m": 235.0, "methodology": "EV/Sales 1.4x"},
        },
        "_engagement_title": "TargetCo M&A diligence",
        "_target_name": "TargetCo Holdings",
        "_firm_name": "Test Firm",
    }


def _citations() -> list[ClaimCitation]:
    return [
        ClaimCitation(
            claim_id="claim_1", text="Revenue trajectory ledger",
            source_title="TargetCo CIM", source_type="firm_library",
        ),
        ClaimCitation(
            claim_id="claim_2", text="EBITDA trajectory ledger",
            source_title="10-K 2023", source_type="sec_filing",
        ),
    ]


@pytest.fixture
def exporter() -> ExcelModelExporter:
    return ExcelModelExporter()


def _all_text(ws: Any) -> str:
    parts: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for cell_value in row:
            if cell_value is not None:
                parts.append(str(cell_value))
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Test 1 — round trip: bytes → reopen → 2 sheets
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_renders_xlsx_round_trip(exporter: ExcelModelExporter) -> None:
    result = await exporter.render(_payload(), _BRANDING, _citations())
    assert result.file_size > 0
    # XLSX is a ZIP — header magic is PK\x03\x04.
    assert result.file_bytes[:4] == b"PK\x03\x04"
    wb = load_workbook(io.BytesIO(result.file_bytes))
    assert wb.sheetnames == ["Cover", "Assumptions"]
    assert result.metadata["mode"] == "m_and_a_diligence"
    assert result.metadata["sheet_sequence"] == ["title", "assumptions"]
    assert result.metadata["sheet_count"] == 2


# ---------------------------------------------------------------------------
# Test 2 — Cover sheet contains the engagement title + firm name
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_cover_sheet_contains_engagement_title(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    text = _all_text(wb["Cover"])
    assert "Test Firm" in text                     # firm-name banner (A1)
    assert "TargetCo M&A diligence" in text         # engagement title (A2)
    assert "TargetCo Holdings" in text              # target (A3)
    assert "Prepared by Test Firm" in text          # A5 line


# ---------------------------------------------------------------------------
# Test 3 — Assumptions input cells use the BLUE-on-YELLOW convention
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_assumptions_sheet_has_color_coded_inputs(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Assumptions"]

    # openpyxl stores fill / font rgb as ARGB (8-char hex) when read
    # back. Strip the leading 2 alpha hex chars so the comparison
    # matches our 6-char RGB constants. (Not a character-set strip —
    # the alpha can be "00" or "FF" depending on construction path.)
    def _norm(rgb: Any) -> str:
        if rgb is None:
            return ""
        s = str(rgb).upper()
        return s[2:] if len(s) == 8 else s

    expected_fill = INPUT_FILL_HEX.upper()
    expected_text = INPUT_TEXT_HEX.upper()
    yellow_inputs = 0
    blue_text_count = 0
    for row in ws.iter_rows():
        for cell in row:
            try:
                fg = cell.fill.fgColor
                if fg is not None and fg.rgb and _norm(fg.rgb) == expected_fill:
                    yellow_inputs += 1
            except Exception:
                pass
            try:
                if cell.font and cell.font.color and cell.font.color.rgb:
                    if _norm(cell.font.color.rgb) == expected_text:
                        blue_text_count += 1
            except Exception:
                pass
    # Spec calls for WACC, terminal growth, tax rate, revenue-growth =
    # at least 4 input cells.
    assert yellow_inputs >= 4, f"expected ≥4 yellow input cells, got {yellow_inputs}"
    assert blue_text_count >= 4, f"expected ≥4 blue-text cells, got {blue_text_count}"


# ---------------------------------------------------------------------------
# Test 4 — Assumptions sheet has WACC / terminal growth / tax rate defaults
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_assumptions_sheet_has_default_wacc_terminal_growth_tax(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Assumptions"]
    rows_by_label: dict[str, Any] = {}
    for row in ws.iter_rows(values_only=False):
        if row and row[0].value is not None:
            rows_by_label[str(row[0].value)] = row
    # Find each parameter and check its default value lives in column B.
    assert "WACC" in rows_by_label
    assert rows_by_label["WACC"][1].value == pytest.approx(0.10)
    assert "Terminal growth rate" in rows_by_label
    assert rows_by_label["Terminal growth rate"][1].value == pytest.approx(0.025)
    assert "Tax rate" in rows_by_label
    assert rows_by_label["Tax rate"][1].value == pytest.approx(0.25)
    # Source/notes column flags defaults.
    for label in ("WACC", "Terminal growth rate", "Tax rate"):
        note = str(rows_by_label[label][3].value or "")
        assert "ASSUMPTION" in note.upper(), (
            f"row {label!r}: source/notes missing ASSUMPTION flag — got {note!r}"
        )


# ---------------------------------------------------------------------------
# Test 5 — payload-derived cells carry citation comments
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_payload_derived_cells_have_citation_comments(
    exporter: ExcelModelExporter,
) -> None:
    result = await exporter.render(_payload(), _BRANDING, _citations())
    wb = load_workbook(io.BytesIO(result.file_bytes))
    ws = wb["Assumptions"]

    cells_with_comments = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment is not None:
                cells_with_comments.append((cell.coordinate, cell.comment.text))
    assert len(cells_with_comments) >= 1, (
        "expected ≥1 cell with a citation comment on the Assumptions sheet"
    )
    # Each comment text leads with [claim_id] then the breadcrumb.
    for coord, text in cells_with_comments:
        assert text.startswith("[claim_"), (
            f"comment on {coord} not in [claim_id] breadcrumb format: {text!r}"
        )
    # Per-deck citation registry surfaces both claim_1 + claim_2.
    citation_ids_in_metadata = result.metadata["cited_claim_ids"]
    assert "claim_1" in citation_ids_in_metadata
    assert "claim_2" in citation_ids_in_metadata


# ---------------------------------------------------------------------------
# Test 6 — write to disk + reopen via load_workbook(filepath) without errors
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_xlsx_opens_without_corruption(
    exporter: ExcelModelExporter, tmp_path: Path,
) -> None:
    result = await exporter.render(_payload(), _BRANDING, _citations())
    fpath = tmp_path / "model.xlsx"
    fpath.write_bytes(result.file_bytes)
    assert fpath.exists()
    assert fpath.stat().st_size == result.file_size
    # Reopen — raises on malformed XLSX.
    wb = load_workbook(str(fpath))
    assert wb.sheetnames == ["Cover", "Assumptions"]
    # File size sanity (no charts / images yet; D1 cap is 50 KB).
    assert result.file_size < 50_000, (
        f"xlsx larger than expected on D1: {result.file_size} bytes (cap 50 KB)"
    )
