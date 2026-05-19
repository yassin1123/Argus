"""Sensitivity sheet — W12/D3 (M&A-specific).

Four 2-axis sensitivity grids:

  Grid 1 — WACC × Terminal Growth → Enterprise Value
  Grid 2 — WACC × Exit Multiple   → Enterprise Value
  Grid 3 — Revenue Growth × EBITDA Margin → Terminal EBITDA
  Grid 4 — DSO × DPO              → Terminal NWC

Per spec hard rule, cells are STATIC pre-computed values rather than
live Excel DATA TABLE features (openpyxl can't reliably write data
tables). Argus computes the values at render time using a pure-Python
DCF that mirrors the DCF sheet's formulas. A note at the top of each
grid tells the consultant "to refresh, re-generate the model in Argus
after editing the underlying assumptions."

The Python DCF is intentionally simple — it mirrors only the cells
the grid touches, not the full cell-for-cell sheet logic.
"""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter

from ..._base import payload_get
from .._refs import NUMBER_FORMATS
from .._styles import (
    HEADING_TEXT_HEX,
    MUTED_TEXT_HEX,
    formula_font,
    heading_font,
    left_align,
    muted_font,
    right_align,
    section_fill,
    style_label,
    thin_border,
)
from ._base import SheetBuilderBase, SheetResult
from ._registry import register_sheet


# Standard sensitivity-grid axes.
_WACC_VALUES =     [0.08, 0.09, 0.10, 0.11, 0.12]
_TG_VALUES =       [0.015, 0.020, 0.025, 0.030, 0.035]
_EXIT_MULT_VALUES = [6.0, 7.0, 8.0, 9.0, 10.0]
_REV_GROWTH_VALUES = [0.03, 0.04, 0.05, 0.06, 0.07]
_EBITDA_MARGIN_VALUES = [0.08, 0.10, 0.12, 0.14, 0.16]
_DSO_VALUES = [30, 40, 50, 60, 70]
_DPO_VALUES = [20, 30, 40, 50, 60]


def _dcf_ev_gordon(
    *,
    fcf_y1: float,
    growth_rate: float,
    years: int,
    wacc: float,
    terminal_growth: float,
) -> float | None:
    """Static DCF helper: assume FCF grows at ``growth_rate`` for
    ``years`` years, then Gordon Growth terminal value at
    ``terminal_growth``, discounted at ``wacc``."""
    if wacc <= terminal_growth:
        return None  # Gordon Growth undefined.
    fcf_stream = [fcf_y1 * ((1 + growth_rate) ** (t - 1)) for t in range(1, years + 1)]
    pv = sum(f / ((1 + wacc) ** t) for t, f in enumerate(fcf_stream, start=1))
    tv = (fcf_stream[-1] * (1 + terminal_growth)) / (wacc - terminal_growth)
    pv_tv = tv / ((1 + wacc) ** years)
    return pv + pv_tv


def _dcf_ev_exit_multiple(
    *,
    fcf_y1: float,
    growth_rate: float,
    years: int,
    wacc: float,
    terminal_ebitda: float,
    exit_multiple: float,
) -> float:
    """Static DCF helper: discount FCF stream + (EBITDA_terminal × exit)
    back at WACC."""
    fcf_stream = [fcf_y1 * ((1 + growth_rate) ** (t - 1)) for t in range(1, years + 1)]
    pv = sum(f / ((1 + wacc) ** t) for t, f in enumerate(fcf_stream, start=1))
    tv = terminal_ebitda * exit_multiple
    pv_tv = tv / ((1 + wacc) ** years)
    return pv + pv_tv


def _latest_value(payload: Any, key: str) -> float:
    """Latest period value from a trajectory; 0.0 fallback."""
    fp = payload_get(payload, "financial_profile", default=None)
    if not isinstance(fp, dict):
        return 0.0
    from ...one_pager_renderer import _coerce_to_list
    traj = fp.get(key) or {}
    if not isinstance(traj, dict):
        return 0.0
    pts = _coerce_to_list(traj.get("points") or [])
    for p in reversed(pts):
        if isinstance(p, dict):
            try:
                return float(p.get("value_gbp_m"))
            except (TypeError, ValueError):
                continue
    return 0.0


def _latest_ebitda_margin(payload: Any) -> float:
    fp = payload_get(payload, "financial_profile", default=None)
    if not isinstance(fp, dict):
        return 0.10
    mp = fp.get("margin_profile") or {}
    raw = str(mp.get("ebitda_margin") or "").rstrip("%").strip()
    try:
        return float(raw) / 100.0
    except (ValueError, TypeError):
        return 0.10


def _projection_years_for(mode: str | None) -> int:
    from .revenue_build import _projection_years_for as _resolve
    return _resolve(mode)


def _write_grid(
    ws: Any,
    *,
    title: str,
    note: str,
    row_start: int,
    row_label: str,
    col_label: str,
    row_values: list[float | int],
    col_values: list[float | int],
    compute_cell: Any,  # callable(row_val, col_val) -> float | None
    row_value_format: str,
    col_value_format: str,
    cell_format: str,
    primary_hex: str,
) -> int:
    """Render one sensitivity grid. Returns the row index after the
    grid (so the caller can place the next grid below it)."""
    # Title band.
    for col in range(1, 8):
        c = ws.cell(row=row_start, column=col)
        c.fill = section_fill()
        c.border = thin_border()
    ws.cell(row=row_start, column=1).value = title
    ws.cell(row=row_start, column=1).font = heading_font(color_hex=primary_hex, size=12)
    ws.cell(row=row_start + 1, column=1).value = note
    ws.cell(row=row_start + 1, column=1).font = muted_font(size=9)

    # Top-left axis labels.
    ws.cell(row=row_start + 3, column=1).value = f"{row_label} ↓ / {col_label} →"
    ws.cell(row=row_start + 3, column=1).font = muted_font(size=10)
    ws.cell(row=row_start + 3, column=1).alignment = left_align()

    # Column header values (col_values along the top).
    for j, cv in enumerate(col_values):
        c = ws.cell(row=row_start + 3, column=2 + j, value=cv)
        c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=10)
        c.alignment = right_align()
        c.number_format = col_value_format
        c.fill = section_fill()
        c.border = thin_border()

    # Row label values (row_values down the left) + cells.
    for i, rv in enumerate(row_values):
        rr = row_start + 4 + i
        rc = ws.cell(row=rr, column=1, value=rv)
        rc.font = heading_font(color_hex=MUTED_TEXT_HEX, size=10)
        rc.alignment = right_align()
        rc.number_format = row_value_format
        rc.fill = section_fill()
        rc.border = thin_border()
        for j, cv in enumerate(col_values):
            cell = ws.cell(row=rr, column=2 + j)
            try:
                value = compute_cell(rv, cv)
            except Exception:
                value = None
            if value is None:
                cell.value = "n/a"
                cell.font = muted_font(size=10)
            else:
                cell.value = round(value, 1)
                cell.font = formula_font(bold=False)
            cell.alignment = right_align()
            cell.number_format = cell_format
            cell.border = thin_border()
    return row_start + 4 + len(row_values)


@register_sheet("sensitivity")
class SensitivitySheet(SheetBuilderBase):
    def build(
        self,
        workbook: Any,
        payload: Any,
        firm_branding: dict[str, Any],
        citations: list[Any],
        cell_registry: Any = None,
    ) -> SheetResult:
        ws = workbook.create_sheet("Sensitivity")

        mode_hint = payload_get(payload, "_mode_hint", default=None)
        explicit_mode = str(payload_get(payload, "mode", default="") or "").strip()
        mode = mode_hint or explicit_mode or "general"
        proj_years = _projection_years_for(mode)

        primary_hex = str(
            (firm_branding or {}).get("primary_color") or f"#{HEADING_TEXT_HEX}"
        ).lstrip("#")

        ws["A1"] = "Sensitivity"
        ws["A1"].font = heading_font(color_hex=primary_hex, size=18)
        ws.row_dimensions[1].height = 28

        # Column widths.
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.column_dimensions["A"].width = 36

        # Static-values note across the top.
        ws["A2"] = (
            "Values computed at generation time. Re-run Argus after editing "
            "Assumptions to refresh."
        )
        ws["A2"].font = muted_font(size=10)
        ws["A2"].alignment = left_align()
        ws.merge_cells("A2:G2")

        # Compute base inputs for the static DCFs.
        base_revenue = _latest_value(payload, "revenue_trajectory")
        base_ebitda = _latest_value(payload, "ebitda_trajectory")
        if base_ebitda <= 0 and base_revenue > 0:
            base_ebitda = base_revenue * _latest_ebitda_margin(payload)
        # FCF Y1 is a simplification — base_ebitda × (1 - tax) × (1 - capex_pct).
        # Tax 0.25, capex 0.05, no NWC.
        base_fcf_y1 = base_ebitda * (1 - 0.25) - base_revenue * 0.05
        base_growth = 0.05
        base_terminal_ebitda = base_ebitda * (1 + base_growth) ** (proj_years - 1)

        # Grid 1 — WACC × Terminal Growth → EV (Gordon Growth)
        next_row = _write_grid(
            ws,
            title="Grid 1 — WACC × Terminal Growth → Enterprise Value (£m)",
            note="Gordon Growth method; FCF grows at base 5% before terminal.",
            row_start=4,
            row_label="WACC",
            col_label="Terminal Growth",
            row_values=_WACC_VALUES,
            col_values=_TG_VALUES,
            compute_cell=lambda wacc, tg: _dcf_ev_gordon(
                fcf_y1=base_fcf_y1, growth_rate=base_growth, years=proj_years,
                wacc=wacc, terminal_growth=tg,
            ),
            row_value_format=NUMBER_FORMATS["percent"],
            col_value_format=NUMBER_FORMATS["percent"],
            cell_format=NUMBER_FORMATS["currency_gbp_m"],
            primary_hex=primary_hex,
        )

        # Grid 2 — WACC × Exit Multiple → EV
        next_row = _write_grid(
            ws,
            title="Grid 2 — WACC × Exit Multiple → Enterprise Value (£m)",
            note="Exit-multiple method; FCF grows at base 5% before terminal.",
            row_start=next_row + 2,
            row_label="WACC",
            col_label="Exit Multiple",
            row_values=_WACC_VALUES,
            col_values=_EXIT_MULT_VALUES,
            compute_cell=lambda wacc, em: _dcf_ev_exit_multiple(
                fcf_y1=base_fcf_y1, growth_rate=base_growth, years=proj_years,
                wacc=wacc, terminal_ebitda=base_terminal_ebitda,
                exit_multiple=em,
            ),
            row_value_format=NUMBER_FORMATS["percent"],
            col_value_format=NUMBER_FORMATS["multiple"],
            cell_format=NUMBER_FORMATS["currency_gbp_m"],
            primary_hex=primary_hex,
        )

        # Grid 3 — Revenue Growth × EBITDA Margin → Terminal EBITDA
        def _terminal_ebitda(growth: float, margin: float) -> float:
            terminal_revenue = base_revenue * ((1 + growth) ** proj_years)
            return terminal_revenue * margin

        next_row = _write_grid(
            ws,
            title="Grid 3 — Revenue Growth × EBITDA Margin → Terminal EBITDA (£m)",
            note=f"Compounds base revenue (£{base_revenue:.0f}m) over {proj_years} years × margin.",
            row_start=next_row + 2,
            row_label="Revenue Growth",
            col_label="EBITDA Margin",
            row_values=_REV_GROWTH_VALUES,
            col_values=_EBITDA_MARGIN_VALUES,
            compute_cell=_terminal_ebitda,
            row_value_format=NUMBER_FORMATS["percent"],
            col_value_format=NUMBER_FORMATS["percent"],
            cell_format=NUMBER_FORMATS["currency_gbp_m"],
            primary_hex=primary_hex,
        )

        # Grid 4 — DSO × DPO → Terminal NWC
        def _terminal_nwc(dso: float, dpo: float) -> float:
            terminal_revenue = base_revenue * ((1 + base_growth) ** proj_years)
            inv_days = 25.0
            return terminal_revenue * (dso + inv_days - dpo) / 365.0

        next_row = _write_grid(
            ws,
            title="Grid 4 — DSO × DPO → Terminal NWC (£m)",
            note="Inventory days held at 25.",
            row_start=next_row + 2,
            row_label="DSO",
            col_label="DPO",
            row_values=_DSO_VALUES,
            col_values=_DPO_VALUES,
            compute_cell=_terminal_nwc,
            row_value_format=NUMBER_FORMATS["integer"],
            col_value_format=NUMBER_FORMATS["integer"],
            cell_format=NUMBER_FORMATS["currency_gbp_m"],
            primary_hex=primary_hex,
        )

        return SheetResult(
            sheet_index=workbook.worksheets.index(ws),
            citation_ids=[],
            cell_count=4 * 25,  # 4 grids × 5×5 cells
        )
