"""Working Capital sheet — W12/D3 (M&A-specific).

The freeform ``payload.financial_profile.working_capital_dynamics``
string ("51-day WC cycle, supplier-funded shape.") can't be reliably
parsed to DSO / DPO / inventory numbers — the writer doesn't enforce
a structured schema for those today. So this sheet uses BLUE-on-YELLOW
input cells with reasonable defaults (consultant-modelling discipline)
and a citation comment on the WC narrative pointing back to the
payload.

Rows:
  - DSO (Days Sales Outstanding)        — input, default 45
  - DPO (Days Payable Outstanding)      — input, default 30
  - Inventory Days                       — input, default 25
  - Net Working Capital                  — formula:
       Revenue × (DSO + Inventory - DPO) / 365
  - ΔNWC                                  — formula: current - prior
                                            (feeds DCF Free Cash Flow)

Registers ``nwc_change_y1`` … ``nwc_change_yN`` so the DCF sheet
chains cleanly. Skipped entirely (not added to the workbook) if
financial_profile is missing — the DCF sheet then shows the
documented "DCF not available" placeholder.
"""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter

from ..._base import payload_get
from .._refs import NUMBER_FORMATS, absolute_ref
from .._styles import (
    HEADING_TEXT_HEX,
    MUTED_TEXT_HEX,
    formula_font,
    heading_font,
    left_align,
    link_font,
    muted_font,
    right_align,
    section_fill,
    style_input_cell,
    style_label,
    thin_border,
)
from ..citation_helpers import add_citation_comment, breadcrumb_for_citation
from ._base import SheetBuilderBase, SheetResult
from ._registry import register_sheet


_DEFAULT_DSO = 45
_DEFAULT_DPO = 30
_DEFAULT_INV = 25
_ASSUMPTION_NOTE = "ASSUMPTION — review before use"


def _projection_years_for(mode: str | None) -> int:
    from .revenue_build import _projection_years_for as _resolve
    return _resolve(mode)


def _historical_periods(payload: Any) -> list[str]:
    fp = payload_get(payload, "financial_profile", default=None)
    if not isinstance(fp, dict):
        return []
    from ...one_pager_renderer import _coerce_to_list

    traj = fp.get("revenue_trajectory") or {}
    if not isinstance(traj, dict):
        return []
    out: list[str] = []
    for p in _coerce_to_list(traj.get("points") or [])[-5:]:
        if isinstance(p, dict) and str(p.get("period") or "").strip():
            out.append(str(p["period"]).strip())
    return out


def _wc_narrative(payload: Any) -> str:
    fp = payload_get(payload, "financial_profile", default=None)
    if not isinstance(fp, dict):
        return ""
    return str(fp.get("working_capital_dynamics") or "").strip()


@register_sheet("working_capital")
class WorkingCapitalSheet(SheetBuilderBase):
    def build(
        self,
        workbook: Any,
        payload: Any,
        firm_branding: dict[str, Any],
        citations: list[Any],
        cell_registry: Any = None,
    ) -> SheetResult:
        ws = workbook.create_sheet("Working Capital")

        mode_hint = payload_get(payload, "_mode_hint", default=None)
        explicit_mode = str(payload_get(payload, "mode", default="") or "").strip()
        mode = mode_hint or explicit_mode or "general"
        proj_years = _projection_years_for(mode)

        primary_hex = str(
            (firm_branding or {}).get("primary_color") or f"#{HEADING_TEXT_HEX}"
        ).lstrip("#")

        ws["A1"] = "Working Capital"
        ws["A1"].font = heading_font(color_hex=primary_hex, size=18)
        ws.row_dimensions[1].height = 28

        historicals = _historical_periods(payload)
        n_hist = len(historicals)

        # Column widths.
        widths = {1: 32}
        for col in range(2, 2 + n_hist + proj_years):
            widths[col] = 14
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # Year header row.
        for col, label in enumerate(["Period"] + historicals + [f"FY+{j}" for j in range(1, proj_years + 1)], start=1):
            c = ws.cell(row=3, column=col, value=label)
            c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
            c.alignment = right_align() if col > 1 else left_align()
            c.fill = section_fill()
            c.border = thin_border()
        ws.row_dimensions[3].height = 22

        # WC narrative blurb (row 5).
        narrative = _wc_narrative(payload)
        if narrative:
            ws["A5"] = narrative
            ws["A5"].font = muted_font(size=10)
            ws["A5"].alignment = left_align()
            ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2 + n_hist + proj_years - 1)

        # Input rows (DSO / DPO / Inventory). Values constant across years
        # (this is the conventional v1 model — assume steady-state days).
        input_rows = {
            "dso": ("DSO (Days Sales Outstanding)", _DEFAULT_DSO),
            "dpo": ("DPO (Days Payable Outstanding)", _DEFAULT_DPO),
            "inv": ("Inventory Days", _DEFAULT_INV),
        }
        days_row: dict[str, int] = {}
        row = 7
        for key, (label, default) in input_rows.items():
            style_label(ws.cell(row=row, column=1), value=label)
            for col in range(2, 2 + n_hist + proj_years):
                cell = ws.cell(row=row, column=col)
                style_input_cell(cell, value=default,
                                 number_format=NUMBER_FORMATS["integer"])
            days_row[key] = row
            row += 1
        cell_count = 3 * (n_hist + proj_years)

        # Computed rows: Net Working Capital + ΔNWC.
        # NWC = Revenue × (DSO + Inv - DPO) / 365
        row_nwc = row
        style_label(ws.cell(row=row_nwc, column=1), value="Net Working Capital", bold=True)
        row += 1
        row_dnwc = row
        style_label(ws.cell(row=row_dnwc, column=1), value="ΔNWC", bold=True)
        row += 1

        # Compute NWC + ΔNWC per column.
        # Historicals: NWC = (DSO + Inv - DPO)/365 × placeholder
        # (we don't have historical revenue on this sheet directly;
        # link to Revenue Build totals via the registry).
        for j in range(n_hist + proj_years):
            col = 2 + j
            col_letter = get_column_letter(col)
            if j < n_hist:
                # Historical period — no revenue link registered for
                # historical years today, so leave NWC + ΔNWC blank
                # (consultant fills if needed).
                ws.cell(row=row_nwc, column=col).value = None
                ws.cell(row=row_dnwc, column=col).value = None
            else:
                year = j - n_hist + 1
                rev_ref = (
                    cell_registry.get(f"revenue_y{year}") if cell_registry is not None
                    else None
                )
                if rev_ref:
                    nwc_formula = (
                        f"={rev_ref}*({col_letter}{days_row['dso']}"
                        f"+{col_letter}{days_row['inv']}"
                        f"-{col_letter}{days_row['dpo']})/365"
                    )
                else:
                    nwc_formula = None
                nwc_cell = ws.cell(row=row_nwc, column=col)
                if nwc_formula:
                    nwc_cell.value = nwc_formula
                    nwc_cell.font = formula_font(bold=True)
                else:
                    nwc_cell.value = "(Revenue Build unavailable)"
                    nwc_cell.font = muted_font()
                nwc_cell.alignment = right_align()
                nwc_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
                nwc_cell.border = thin_border()

                # ΔNWC = NWC_this - NWC_prior. For year 1 there's no
                # prior on this sheet — fall back to 0 (or skip).
                dnwc_cell = ws.cell(row=row_dnwc, column=col)
                if year == 1:
                    dnwc_cell.value = f"={col_letter}{row_nwc}*0"
                else:
                    prev_letter = get_column_letter(col - 1)
                    dnwc_cell.value = f"={col_letter}{row_nwc}-{prev_letter}{row_nwc}"
                dnwc_cell.font = formula_font()
                dnwc_cell.alignment = right_align()
                dnwc_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
                dnwc_cell.border = thin_border()

                # Register ΔNWC for the DCF sheet.
                if cell_registry is not None:
                    cell_registry.set(
                        f"nwc_change_y{year}",
                        "Working Capital",
                        f"{col_letter}{row_dnwc}",
                    )

        return SheetResult(
            sheet_index=workbook.worksheets.index(ws),
            citation_ids=[],
            cell_count=cell_count,
        )
