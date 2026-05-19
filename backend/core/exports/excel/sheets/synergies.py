"""Synergies sheet — W12/D3 (M&A-specific).

Three tables stacked vertically:
  Revenue Synergies   (payload.synergy_estimate.revenue_synergies)
  Cost Synergies      (payload.synergy_estimate.cost_synergies)
  Dis-synergies       (payload.synergy_estimate.dis_synergies — values
                       rendered as NEGATIVE on the NPV calculation so
                       the sheet reads honestly even when the writer
                       stored them as positives.)

Each row carries:
  type | description | magnitude (£m) | timing (months) | confidence
        | basis citations (Excel comment)

NPV section at the bottom: for each synergy, assume linear realization
over ``timing_months / 12`` years, discount cash flows at WACC, sum
to NPV. NPV row formulas reference Assumptions!wacc so the consultant
can tune WACC and see NPVs recalc.

Optional bar chart shows realized synergy magnitude per projection
year. Skipped when no synergies exist.
"""

from __future__ import annotations

from typing import Any

from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from ..._base import payload_get
from ...one_pager_renderer import _coerce_to_list
from .._refs import NUMBER_FORMATS, absolute_ref
from .._styles import (
    HEADING_TEXT_HEX,
    MUTED_TEXT_HEX,
    formula_font,
    heading_font,
    left_align,
    muted_font,
    right_align,
    section_fill,
    style_input_cell,
    style_label,
    thin_border,
)
from ..citation_helpers import add_citation_comment, breadcrumb_for_citation
from ._base import SheetBuilderBase, SheetResult
from ._registry import register_sheet


def _build_citation_index(citations: list[Any]) -> dict[str, str]:
    out: dict[str, str] = {}
    for c in citations or []:
        cid = (getattr(c, "claim_id", "") or "").strip()
        if cid and cid not in out:
            out[cid] = breadcrumb_for_citation(c)
    return out


def _projection_years_for(mode: str | None) -> int:
    from .revenue_build import _projection_years_for as _resolve
    return _resolve(mode)


@register_sheet("synergies")
class SynergiesSheet(SheetBuilderBase):
    def build(
        self,
        workbook: Any,
        payload: Any,
        firm_branding: dict[str, Any],
        citations: list[Any],
        cell_registry: Any = None,
    ) -> SheetResult:
        ws = workbook.create_sheet("Synergies")

        mode_hint = payload_get(payload, "_mode_hint", default=None)
        explicit_mode = str(payload_get(payload, "mode", default="") or "").strip()
        mode = mode_hint or explicit_mode or "general"
        proj_years = _projection_years_for(mode)

        primary_hex = str(
            (firm_branding or {}).get("primary_color") or f"#{HEADING_TEXT_HEX}"
        ).lstrip("#")

        ws["A1"] = "Synergies"
        ws["A1"].font = heading_font(color_hex=primary_hex, size=18)
        ws.row_dimensions[1].height = 28

        # Column widths.
        widths = {1: 30, 2: 36, 3: 12, 4: 12, 5: 12, 6: 12}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        for col in range(7, 7 + proj_years):
            ws.column_dimensions[get_column_letter(col)].width = 12

        synergy_estimate = payload_get(payload, "synergy_estimate", default={}) or {}
        if not isinstance(synergy_estimate, dict):
            synergy_estimate = {}

        sections = [
            ("Revenue synergies", "revenue_synergies", +1),
            ("Cost synergies", "cost_synergies", +1),
            ("Dis-synergies", "dis_synergies", -1),
        ]

        cited: list[str] = []
        cite_index = _build_citation_index(citations)
        all_synergy_rows: list[tuple[int, int]] = []  # (row, sign)
        cell_count = 0

        row = 3
        wacc_ref = (
            cell_registry.get("wacc") if cell_registry is not None
            else absolute_ref("Assumptions", "B14")
        )

        # Common column headers.
        def _write_headers(header_row: int) -> None:
            headers = ["Type", "Description", "Magnitude (£m)", "Timing (mo)", "Confidence"]
            for c_idx, label in enumerate(headers, start=1):
                c = ws.cell(row=header_row, column=c_idx, value=label)
                c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
                c.alignment = left_align() if c_idx <= 2 else right_align()
                c.fill = section_fill(); c.border = thin_border()
            # NPV column header.
            c = ws.cell(row=header_row, column=6, value="NPV (£m)")
            c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
            c.alignment = right_align()
            c.fill = section_fill(); c.border = thin_border()

        for section_title, key, sign in sections:
            entries = _coerce_to_list(synergy_estimate.get(key) or [])
            entries = [e for e in entries if isinstance(e, dict)]

            # Section title band.
            for col in range(1, 7):
                c = ws.cell(row=row, column=col)
                c.fill = section_fill(); c.border = thin_border()
            ws.cell(row=row, column=1).value = section_title
            ws.cell(row=row, column=1).font = heading_font(
                color_hex=primary_hex, size=12,
            )
            row += 1
            _write_headers(row)
            row += 1

            if not entries:
                ws.cell(row=row, column=1).value = (
                    f"(no {section_title.lower()} cited)"
                )
                ws.cell(row=row, column=1).font = muted_font()
                ws.cell(row=row, column=1).alignment = left_align()
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                row += 2  # spacer
                continue

            for s in entries[:10]:
                synergy_type = str(s.get("type") or "").strip() or "—"
                description = synergy_type  # ``type`` doubles as the prose label
                magnitude = s.get("magnitude_gbp_m")
                try:
                    mag_val = float(magnitude)
                except (TypeError, ValueError):
                    mag_val = 0.0
                signed_mag = sign * abs(mag_val)
                timing = s.get("timing_months")
                try:
                    timing_val = int(timing)
                except (TypeError, ValueError):
                    timing_val = 24
                confidence = str(s.get("confidence") or "medium").strip()
                basis = list(s.get("basis_citations") or [])

                style_label(ws.cell(row=row, column=1), value=synergy_type[:80])
                style_label(ws.cell(row=row, column=2), value=description[:120])

                mag_cell = ws.cell(row=row, column=3)
                mag_cell.value = signed_mag
                mag_cell.font = formula_font()
                mag_cell.alignment = right_align()
                mag_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
                mag_cell.border = thin_border()
                # Cite the magnitude cell.
                if basis:
                    cid = str(basis[0])
                    breadcrumb = cite_index.get(cid) or cid
                    add_citation_comment(mag_cell, claim_id=cid, citation_text=breadcrumb)
                    if cid not in cited:
                        cited.append(cid)

                timing_cell = ws.cell(row=row, column=4)
                style_input_cell(timing_cell, value=timing_val,
                                 number_format=NUMBER_FORMATS["integer"])

                conf_cell = ws.cell(row=row, column=5)
                style_input_cell(conf_cell, value=confidence, number_format="@")

                # NPV: assume linear realization over (timing/12) years,
                # discounted at WACC. Approximate as a single
                # mid-period bullet payment for simplicity in v1.
                # NPV ≈ Magnitude * 1/(1+WACC)^((timing/12)/2)
                # Excel formula form for transparency:
                npv_cell = ws.cell(row=row, column=6)
                # Use C, D refs so an Excel-savvy consultant can read the math.
                col_c = "C"; col_d = "D"
                npv_cell.value = (
                    f"={col_c}{row}/(1+{wacc_ref})^(({col_d}{row}/12)/2)"
                )
                npv_cell.font = formula_font(bold=True)
                npv_cell.alignment = right_align()
                npv_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
                npv_cell.border = thin_border()
                all_synergy_rows.append((row, sign))

                row += 1
                cell_count += 5

            row += 1  # gap

        # ---- Summary row: total NPV across synergies ----
        if all_synergy_rows:
            for col in range(1, 7):
                c = ws.cell(row=row, column=col)
                c.fill = section_fill(); c.border = thin_border()
            ws.cell(row=row, column=1).value = "Total NPV (all synergies)"
            ws.cell(row=row, column=1).font = heading_font(
                color_hex=primary_hex, size=12,
            )
            npv_cells = [f"F{r}" for r, _ in all_synergy_rows]
            sum_cell = ws.cell(row=row, column=6)
            sum_cell.value = f"=SUM({','.join(npv_cells)})"
            sum_cell.font = formula_font(bold=True)
            sum_cell.alignment = right_align()
            sum_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
            sum_cell.border = thin_border()
            row += 2
            cell_count += 1

            # ---- Realization timeline (single bar per synergy) ----
            ws.cell(row=row, column=1).value = "Realization timeline"
            ws.cell(row=row, column=1).font = heading_font(
                color_hex=primary_hex, size=12,
            )
            row += 1
            ws.cell(row=row, column=1).value = "Synergy"
            ws.cell(row=row, column=1).font = muted_font()
            for j in range(1, proj_years + 1):
                c = ws.cell(row=row, column=1 + j, value=f"FY+{j}")
                c.font = muted_font()
                c.alignment = right_align()
                c.fill = section_fill(); c.border = thin_border()
            row += 1

            # Linear ramp: realized magnitude in year j =
            # mag × min(1, j × 12 / timing_months).
            chart_data_top = row
            for syn_row, syn_sign in all_synergy_rows:
                ws.cell(row=row, column=1).value = ws.cell(row=syn_row, column=1).value
                ws.cell(row=row, column=1).font = muted_font(size=10)
                for j in range(1, proj_years + 1):
                    cell = ws.cell(row=row, column=1 + j)
                    cell.value = (
                        f"=C{syn_row}*MIN(1,{j}*12/D{syn_row})"
                    )
                    cell.font = formula_font()
                    cell.alignment = right_align()
                    cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
                    cell.border = thin_border()
                row += 1
            chart_data_bottom = row - 1

            # Optional bar chart.
            try:
                chart = BarChart()
                chart.type = "col"
                chart.title = "Synergy realization by year"
                chart.style = 11
                chart.y_axis.title = "£m"
                chart.x_axis.title = "Year"
                data_ref = Reference(
                    ws,
                    min_col=2,
                    min_row=chart_data_top - 1,  # header row
                    max_col=1 + proj_years,
                    max_row=chart_data_bottom,
                )
                cats_ref = Reference(
                    ws,
                    min_col=1, min_row=chart_data_top,
                    max_col=1, max_row=chart_data_bottom,
                )
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                ws.add_chart(chart, f"H{chart_data_top - 2}")
            except Exception:
                # Chart construction is best-effort — never block the
                # sheet on chart-API quirks.
                pass

        return SheetResult(
            sheet_index=workbook.worksheets.index(ws),
            citation_ids=cited,
            cell_count=cell_count,
        )
