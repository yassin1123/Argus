"""Cover sheet — W12/D1.

Sheet name: "Cover" (per spec).

Layout (single column for readability):
  A1: firm name (24pt bold, firm primary colour)
  A2: engagement title
  A3: target / subject
  A4: generation date
  A5: "Prepared by {firm.name}"
  A7-A12: model overview — which sheets are included, the input
          colour convention, where to start editing.

No citations / no comments on this sheet — it's the consultant's
landing page, not a derived-data sheet.
"""

from __future__ import annotations

import asyncio
import io
import logging
from datetime import datetime, timezone
from typing import Any

from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter

from ..._base import payload_get
from ...asset_cache import _cache_fresh, _cache_path, fetch_and_cache_logo
from .._styles import (
    DEFAULT_FONT,
    HEADING_TEXT_HEX,
    MUTED_TEXT_HEX,
    heading_font,
    left_align,
    muted_font,
    style_label,
)
from ._base import SheetBuilderBase, SheetResult
from ._registry import register_sheet

logger = logging.getLogger(__name__)


def _resolve_logo_sync(firm_id: Any, logo_url: str) -> bytes | None:
    """Sync wrapper around the async asset cache. Mirrors the W11/D4
    deck title slide helper:
      - outside a running loop: drive the coroutine via asyncio.run
      - inside a running loop: read the cache file directly so we
        never block the rendering coroutine.
    """
    if not logo_url:
        return None
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        loop = None
    if loop is not None and loop.is_running():
        p = _cache_path(firm_id)
        if _cache_fresh(p):
            try:
                return p.read_bytes()
            except OSError:
                return None
        return None
    try:
        return asyncio.run(fetch_and_cache_logo(firm_id, logo_url))
    except Exception as e:  # noqa: BLE001
        logger.info("xlsx Cover logo resolve failed (%s) — falling back to text", e)
        return None


@register_sheet("title")
class TitleSheet(SheetBuilderBase):
    def build(
        self,
        workbook: Any,
        payload: Any,
        firm_branding: dict[str, Any],
        citations: list[Any],
        cell_registry: Any = None,
    ) -> SheetResult:
        ws = workbook.create_sheet("Cover")

        # Column widths — wide first column for the firm-name banner +
        # the model-overview narrative.
        ws.column_dimensions[get_column_letter(1)].width = 56

        firm_name = (
            (firm_branding or {}).get("_firm_name")
            or payload_get(payload, "_firm_name", default="Argus")
            or "Argus"
        )
        primary_hex = str(
            (firm_branding or {}).get("primary_color") or f"#{HEADING_TEXT_HEX}"
        ).lstrip("#")

        # A1: large firm-name banner in firm primary colour.
        ws["A1"] = firm_name
        ws["A1"].font = heading_font(color_hex=primary_hex, size=24)
        ws["A1"].alignment = left_align()
        ws.row_dimensions[1].height = 36

        # Logo embed (top-right, columns C-D, row 1-3) — reuses the
        # W11/D4 deck asset cache so a single fetch + resize +
        # 24h disk TTL covers HTML / PDF / PPTX / XLSX. Falls back
        # silently to the firm-name text banner when the URL is
        # missing / unreachable / undecodable.
        logo_url = str((firm_branding or {}).get("logo_url") or "").strip()
        firm_id = (firm_branding or {}).get("_firm_id") or firm_name
        if logo_url:
            logo_bytes = _resolve_logo_sync(firm_id, logo_url)
            if logo_bytes:
                try:
                    img = OpenpyxlImage(io.BytesIO(logo_bytes))
                    # Constrain visible size — Pillow resize already
                    # caps source at 300px wide, but openpyxl's
                    # OOXML image cell size also needs setting.
                    img.height = min(img.height, 80)
                    img.width = min(img.width, 200)
                    ws.add_image(img, "C1")
                except Exception as e:  # noqa: BLE001
                    logger.info("xlsx Cover logo embed failed (%s)", e)

        # A2-A5: engagement metadata block.
        engagement_title = str(
            payload_get(payload, "_engagement_title", default="Argus engagement")
        )
        target_name = str(payload_get(payload, "_target_name", default="")) or "—"
        date_str = datetime.now(tz=timezone.utc).strftime("%B %Y")

        style_label(ws["A2"], value=engagement_title, bold=True)
        ws["A2"].font = heading_font(color_hex=HEADING_TEXT_HEX, size=16)

        style_label(ws["A3"], value=f"Target: {target_name}")
        style_label(ws["A4"], value=f"Generated: {date_str}")
        style_label(ws["A5"], value=f"Prepared by {firm_name}")
        ws["A5"].font = muted_font()

        # A7-A12: model overview.
        ws["A7"] = "Model overview"
        ws["A7"].font = heading_font(color_hex=HEADING_TEXT_HEX, size=14)
        ws.row_dimensions[7].height = 22

        overview_lines = [
            "This workbook is a starting financial model for the engagement above.",
            "Sheets included (Day 1): Cover · Assumptions.",
            "Editable input cells are formatted with BLUE text on a YELLOW fill —"
            " those are the only values you should change.",
            "Cells with BLACK text are formulas; do not overwrite them.",
            "Hover any cell with a red comment indicator to see the source"
            " breadcrumb that grounds the value.",
            "Days 2-3 of Week 12 add Revenue Build, Cost Build, DCF and Comparables.",
        ]
        for i, text in enumerate(overview_lines, start=8):
            ws[f"A{i}"] = text
            ws[f"A{i}"].font = muted_font(size=11)
            ws[f"A{i}"].alignment = left_align()
            ws.row_dimensions[i].height = 20

        return SheetResult(
            sheet_index=workbook.worksheets.index(ws),
            citation_ids=[],
            cell_count=4 + 1 + len(overview_lines),
            skip_branding_header=True,  # Cover owns its own row-1 banner
        )
