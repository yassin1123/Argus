"""Revenue Build sheet — W12/D2.

Mode-aware revenue projection. Both M&A and growth_strategy get this
sheet, but the projection horizon and segment depth differ:

  M&A:        5 projection years, segment-level detail (rows per
              segment from payload.target_overview.segments), total
              row at the bottom = SUM of segments per column.
  growth:     3 projection years, single-line revenue projection.
  general:    3 projection years, single-line.

Structure on the sheet:
  Row 1:     Title (firm-coloured banner)
  Row 3:     Year header row (FY-3, FY-2, FY-1, FY+0, FY+1, FY+2, ...)
              with HISTORICAL columns as values + citation comments,
              PROJECTION columns as formulas referencing Assumptions.
  Rows 5..N: One row per segment (M&A) or single "Revenue" row
              (others). Historicals are values; projections are
              ``=PREV * (1 + Assumptions!$B$X)`` formulas where the
              growth rate cell is resolved via the CellRegistry.
  Row N+1:   Total revenue (M&A only): SUM of segments per column.

Fallback: when payload.financial_profile.revenue_trajectory is empty,
historical columns render the documented "n/a — financial trajectory
not produced for this engagement" placeholder, and projection columns
render BLUE input cells with a 0 default so a consultant can fill
them in manually.

Hard rules:
  - Historicals are values + citations; projections are formulas
    (consulting-modelling discipline).
  - Projection years never exceed 5 (terminal value handles beyond).
  - No auto-extrapolation of growth from historical CAGR.
"""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter

from ..._base import payload_get
from ...one_pager_renderer import _coerce_to_list
from .._refs import NUMBER_FORMATS, absolute_ref, cell_ref
from .._styles import (
    HEADING_TEXT_HEX,
    MUTED_TEXT_HEX,
    formula_font,
    heading_font,
    input_fill,
    input_font,
    left_align,
    muted_font,
    right_align,
    section_fill,
    style_input_cell,
    style_label,
    thin_border,
)
from ..citation_helpers import add_citation_comment, breadcrumb_for_citation
from ._base import SheetBuilderBase, SheetResult
from ._registry import register_sheet


# Projection horizons per mode. Spec hard rule: cap at 5.
_PROJECTION_YEARS_BY_MODE: dict[str, int] = {
    "m_and_a_diligence": 5,
    "growth_strategy":   3,
    "boutique_pricing_review": 3,
    "market_entry":      3,
    "general":           3,
}
_MAX_PROJECTION_YEARS = 5

# How many historical points to show on the sheet. We rely on the
# writer's trajectory length but cap at 5 so the sheet stays readable.
_MAX_HISTORICAL_POINTS = 5


def _projection_years_for(mode: str | None) -> int:
    return min(
        _PROJECTION_YEARS_BY_MODE.get((mode or "general").strip() or "general", 3),
        _MAX_PROJECTION_YEARS,
    )


def _historical_points(payload: Any) -> list[tuple[str, float, str]]:
    """Return ``[(period, value, source_citation), ...]`` from
    ``payload.financial_profile.revenue_trajectory.points``. Empty
    list when the trajectory is missing or malformed."""
    fp = payload_get(payload, "financial_profile", default=None)
    if not isinstance(fp, dict):
        return []
    traj = fp.get("revenue_trajectory") or {}
    if not isinstance(traj, dict):
        return []
    pts = _coerce_to_list(traj.get("points") or [])
    out: list[tuple[str, float, str]] = []
    for p in pts:
        if not isinstance(p, dict):
            continue
        try:
            v = float(p.get("value_gbp_m"))
        except (TypeError, ValueError):
            continue
        period = str(p.get("period") or "").strip()
        cid = str(p.get("source_citation") or "").strip()
        if period:
            out.append((period, v, cid))
    return out[-_MAX_HISTORICAL_POINTS:]


def _segments(payload: Any) -> list[dict[str, Any]]:
    """Pull segment list from target_overview.segments. Used by M&A
    mode for per-segment projection rows."""
    to = payload_get(payload, "target_overview", default=None)
    if not isinstance(to, dict):
        return []
    segs = _coerce_to_list(to.get("segments") or [])
    out: list[dict[str, Any]] = []
    for s in segs:
        if isinstance(s, dict) and str(s.get("name") or "").strip():
            out.append(s)
    return out


def _build_citation_index(citations: list[Any]) -> dict[str, str]:
    out: dict[str, str] = {}
    for c in citations or []:
        cid = (getattr(c, "claim_id", "") or "").strip()
        if cid and cid not in out:
            out[cid] = breadcrumb_for_citation(c)
    return out


def _is_m_and_a(payload: Any, mode_hint: str | None) -> bool:
    """Same heuristic as one_pager_renderer._detect_mode but simplified
    for "do we expand segments?"."""
    if mode_hint and "m_and_a" in mode_hint:
        return True
    explicit = str(payload_get(payload, "mode", default="") or "").lower()
    if "m_and_a" in explicit:
        return True
    if payload_get(payload, "target_overview", default=None) is not None:
        # Heuristic from the renderer.
        if payload_get(payload, "valuation_range", default=None) is not None:
            return True
    return False


@register_sheet("revenue_build")
class RevenueBuildSheet(SheetBuilderBase):
    def build(
        self,
        workbook: Any,
        payload: Any,
        firm_branding: dict[str, Any],
        citations: list[Any],
        cell_registry: Any = None,
    ) -> SheetResult:
        ws = workbook.create_sheet("Revenue Build")

        # Resolve mode + horizon.
        mode_hint = payload_get(payload, "_mode_hint", default=None)
        explicit_mode = str(payload_get(payload, "mode", default="") or "").strip()
        mode = mode_hint or explicit_mode or "general"
        proj_years = _projection_years_for(mode)
        is_m_and_a = _is_m_and_a(payload, mode_hint)

        # ---- Header ----
        primary_hex = str(
            (firm_branding or {}).get("primary_color") or f"#{HEADING_TEXT_HEX}"
        ).lstrip("#")
        ws["A1"] = "Revenue Build"
        ws["A1"].font = heading_font(color_hex=primary_hex, size=18)
        ws.row_dimensions[1].height = 28

        # ---- Year header row (row 3) ----
        # Column A holds the row label (segment name / "Total");
        # historicals begin at column B, projections follow.
        historicals = _historical_points(payload)
        n_hist = len(historicals)
        # Reserve at most 4 historical columns + projection_years
        # projection columns (so M&A maxes at 9 columns wide,
        # growth at 7).
        # Anchor "FY+0" at the most-recent historical period; the
        # remaining historicals fall back as FY-1, FY-2, FY-3.
        col_widths = {1: 32}
        for col in range(2, 2 + n_hist + proj_years):
            col_widths[col] = 14
        for col, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # Year labels.
        ws.cell(row=3, column=1).value = "Period"
        ws.cell(row=3, column=1).font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
        ws.cell(row=3, column=1).fill = section_fill()
        ws.cell(row=3, column=1).border = thin_border()
        ws.cell(row=3, column=1).alignment = left_align()
        for i, (period, _, _) in enumerate(historicals):
            c = ws.cell(row=3, column=2 + i, value=period)
            c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
            c.alignment = right_align()
            c.fill = section_fill()
            c.border = thin_border()
        for j in range(1, proj_years + 1):
            c = ws.cell(row=3, column=2 + n_hist + j - 1, value=f"FY+{j}")
            c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
            c.alignment = right_align()
            c.fill = section_fill()
            c.border = thin_border()
        ws.row_dimensions[3].height = 22

        cited: list[str] = []
        cell_count = 0
        cite_index = _build_citation_index(citations)

        # ---- Body rows ----
        body_row_start = 5

        if not historicals:
            # FALLBACK: render the placeholder rows.
            _render_fallback(
                ws, body_row_start, proj_years,
                cell_count=lambda: cell_count,
            )
            return SheetResult(
                sheet_index=workbook.worksheets.index(ws),
                citation_ids=[],
                cell_count=2 + proj_years,  # 1 label + N input cells
            )

        # Decide whether to render per-segment rows.
        segments_to_render: list[dict[str, Any]] = []
        if is_m_and_a:
            segs = _segments(payload)
            if segs:
                segments_to_render = segs[:8]  # readability cap

        if segments_to_render:
            # Per-segment rows. Each segment's historical value comes
            # from segment.revenue_pct × total historical revenue.
            # Projection columns reference the previous-year cell in
            # the same row × Assumptions growth rate (single global
            # growth assumption used across segments today; segment-
            # specific growth is W12/D3 polish).
            seg_rows: list[int] = []
            for s_idx, seg in enumerate(segments_to_render):
                row = body_row_start + s_idx
                seg_rows.append(row)
                seg_name = str(seg.get("name") or f"Segment {s_idx+1}").strip()
                style_label(ws.cell(row=row, column=1), value=seg_name)
                try:
                    pct = float(seg.get("revenue_pct"))
                except (TypeError, ValueError):
                    pct = None
                # Historical cells: pct × total revenue.
                for i, (_, value, cid) in enumerate(historicals):
                    cell = ws.cell(row=row, column=2 + i)
                    if pct is not None:
                        cell.value = round(value * pct / 100.0, 1)
                    else:
                        cell.value = None
                    cell.font = formula_font()
                    cell.alignment = right_align()
                    cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
                    cell.border = thin_border()
                    if cid:
                        breadcrumb = cite_index.get(cid) or cid
                        add_citation_comment(cell, claim_id=cid,
                                             citation_text=breadcrumb)
                        if cid not in cited:
                            cited.append(cid)
                # Projection cells: =PREV * (1 + Assumptions!growth_yN)
                for j in range(1, proj_years + 1):
                    col = 2 + n_hist + j - 1
                    prev_col_letter = get_column_letter(col - 1)
                    cell = ws.cell(row=row, column=col)
                    growth_ref = (
                        cell_registry.get(f"revenue_growth_y{j}")
                        if cell_registry is not None
                        else absolute_ref("Assumptions", f"B{14 + j}")
                    )
                    cell.value = (
                        f"={prev_col_letter}{row}*(1+{growth_ref})"
                    )
                    cell.font = formula_font()
                    cell.alignment = right_align()
                    cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
                    cell.border = thin_border()
                cell_count += n_hist + proj_years

            # Total row at the bottom — SUM(segment_rows) per column.
            total_row = seg_rows[-1] + 1
            ws.cell(row=total_row, column=1).value = "Total revenue"
            ws.cell(row=total_row, column=1).font = heading_font(
                color_hex=primary_hex, size=12,
            )
            for j in range(n_hist + proj_years):
                col = 2 + j
                col_letter = get_column_letter(col)
                first = seg_rows[0]
                last = seg_rows[-1]
                cell = ws.cell(row=total_row, column=col)
                cell.value = f"=SUM({col_letter}{first}:{col_letter}{last})"
                cell.font = formula_font(bold=True)
                cell.alignment = right_align()
                cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
                cell.border = thin_border()
                # Register the Total row coords so Cost Build can
                # reference Y+1..Y+N revenue cells cleanly.
                if j >= n_hist and cell_registry is not None:
                    year = j - n_hist + 1
                    cell_registry.set(
                        f"revenue_y{year}",
                        "Revenue Build",
                        f"{col_letter}{total_row}",
                    )

        else:
            # Single-line revenue row.
            row = body_row_start
            style_label(ws.cell(row=row, column=1), value="Revenue", bold=True)
            for i, (_, value, cid) in enumerate(historicals):
                cell = ws.cell(row=row, column=2 + i)
                cell.value = round(value, 1)
                cell.font = formula_font(bold=False)
                cell.alignment = right_align()
                cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
                cell.border = thin_border()
                if cid:
                    breadcrumb = cite_index.get(cid) or cid
                    add_citation_comment(cell, claim_id=cid, citation_text=breadcrumb)
                    if cid not in cited:
                        cited.append(cid)
            for j in range(1, proj_years + 1):
                col = 2 + n_hist + j - 1
                prev_col_letter = get_column_letter(col - 1)
                cell = ws.cell(row=row, column=col)
                growth_ref = (
                    cell_registry.get(f"revenue_growth_y{j}")
                    if cell_registry is not None
                    else absolute_ref("Assumptions", f"B{14 + j}")
                )
                cell.value = f"={prev_col_letter}{row}*(1+{growth_ref})"
                cell.font = formula_font()
                cell.alignment = right_align()
                cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
                cell.border = thin_border()
                # Register Y+N revenue ref so Cost Build can chain.
                if cell_registry is not None:
                    cell_registry.set(
                        f"revenue_y{j}",
                        "Revenue Build",
                        f"{get_column_letter(col)}{row}",
                    )
            cell_count += n_hist + proj_years

        return SheetResult(
            sheet_index=workbook.worksheets.index(ws),
            citation_ids=cited,
            cell_count=cell_count,
        )


# ----------------------------------------------------------------------------
# Fallback path — no financial_profile.revenue_trajectory on payload
# ----------------------------------------------------------------------------


def _render_fallback(
    ws: Any, body_row_start: int, proj_years: int, cell_count: Any,
) -> None:
    """When the payload doesn't carry a revenue trajectory, render a
    single revenue row with placeholder text in historical columns
    and BLUE input cells in projection columns so the consultant can
    fill in manually. The cell carries a comment naming the gap."""
    row = body_row_start

    style_label(ws.cell(row=row, column=1), value="Revenue", bold=True)

    # Historical columns: placeholder string spanning the entire
    # historical region. (Only 1 column to write; if more historicals
    # were expected we'd extend.)
    note_cell = ws.cell(row=row, column=2)
    note_cell.value = (
        "n/a — financial trajectory not produced for this engagement"
    )
    note_cell.font = muted_font()
    note_cell.alignment = left_align()
    note_cell.border = thin_border()
    add_citation_comment(
        note_cell,
        claim_id="payload-gap",
        citation_text="Source data not produced; consultant input required.",
    )

    # Projection columns: blue-on-yellow input cells defaulting to 0.
    for j in range(1, proj_years + 1):
        col = 3 + j - 1  # historical occupied col 2 only
        cell = ws.cell(row=row, column=col)
        style_input_cell(cell, value=0.0,
                         number_format=NUMBER_FORMATS["currency_gbp_m"])
