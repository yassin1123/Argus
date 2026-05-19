"""DCF sheet — W12/D3 (M&A-specific).

Free Cash Flow projection → terminal value (Gordon Growth + Exit
Multiple methods side-by-side) → Enterprise Value → Equity Value.

Row layout:
  1   Title banner
  3   Year header row (FY+1 … FY+N)
  5   Revenue (link to Revenue Build)
  6   EBITDA (link to Cost Build)
  7   Less: Tax (= EBITDA × Assumptions!tax_rate)
  8   Plus: D&A (BLUE input — % of revenue, default 4%)
  9   Less: Capex (BLUE input — % of revenue, default sourced from
      payload.financial_profile.capex_intensity when parseable, else 5%)
  10  Less: ΔNWC (link to Working Capital sheet)
  11  Free Cash Flow (= EBITDA - Tax + D&A - Capex - ΔNWC)
  13  Discount factor (= 1 / (1+WACC)^year)
  14  PV of FCF (= FCF × discount_factor)

  17  Terminal value (Gordon Growth)
        = FCF_terminal × (1 + g) / (WACC - g)
        discounted back at WACC^N
  18  Terminal value (Exit Multiple)
        = EBITDA_terminal × exit_multiple
        discounted back at WACC^N
  19  Selected terminal value (= average of both)

  21  Enterprise Value (= SUM(PV of FCF) + selected TV PV)
  22  Less: Net Debt (BLUE input — payload.debt_structure can't be
       reliably parsed; consultant fills in)
  23  Equity Value (= EV - Net Debt)

Hard-rule compliance:
  - Net Debt is a BLUE assumption cell with a note pointing back at
    the debt_structure narrative (spec hard rule — freeform text
    can't be auto-parsed).
  - WACC < terminal growth would divide by zero on the Gordon
    formula; we still write the formula (Excel will surface
    #DIV/0! to the consultant, which is preferable to a misleading
    silent override).

Registers ``fcf_y1`` … ``fcf_yN`` and ``enterprise_value`` so the
Sensitivity sheet can chain.
"""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter

from ..._base import payload_get
from .._refs import NUMBER_FORMATS, absolute_ref
from .._styles import (
    HEADING_TEXT_HEX,
    MUTED_TEXT_HEX,
    formula_font,
    heading_font,
    left_align,
    link_font,
    muted_font,
    right_align,
    section_fill,
    style_input_cell,
    style_label,
    thin_border,
)
from ._base import SheetBuilderBase, SheetResult
from ._registry import register_sheet


_DEFAULT_DNA_PCT = 0.04
_DEFAULT_CAPEX_PCT = 0.05
_DEFAULT_EXIT_MULTIPLE = 8.0
_DEFAULT_NET_DEBT = 0.0
_ASSUMPTION_NOTE = "ASSUMPTION — review before use"


def _projection_years_for(mode: str | None) -> int:
    from .revenue_build import _projection_years_for as _resolve
    return _resolve(mode)


def _parse_capex_intensity(payload: Any) -> float:
    """Pull a capex/revenue % from
    payload.financial_profile.capex_intensity when parseable.

    Handles strings like ``"4.5% of revenue"`` / ``"Maintenance vs
    growth capex; capex/revenue 5%"`` by extracting the first
    percentage-like number.
    """
    import re
    fp = payload_get(payload, "financial_profile", default=None)
    if not isinstance(fp, dict):
        return _DEFAULT_CAPEX_PCT
    raw = str(fp.get("capex_intensity") or "")
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", raw)
    if not m:
        return _DEFAULT_CAPEX_PCT
    try:
        return float(m.group(1)) / 100.0
    except ValueError:
        return _DEFAULT_CAPEX_PCT


@register_sheet("dcf")
class DCFSheet(SheetBuilderBase):
    def build(
        self,
        workbook: Any,
        payload: Any,
        firm_branding: dict[str, Any],
        citations: list[Any],
        cell_registry: Any = None,
    ) -> SheetResult:
        ws = workbook.create_sheet("DCF")

        mode_hint = payload_get(payload, "_mode_hint", default=None)
        explicit_mode = str(payload_get(payload, "mode", default="") or "").strip()
        mode = mode_hint or explicit_mode or "general"
        proj_years = _projection_years_for(mode)

        primary_hex = str(
            (firm_branding or {}).get("primary_color") or f"#{HEADING_TEXT_HEX}"
        ).lstrip("#")

        ws["A1"] = "DCF — Discounted Cash Flow"
        ws["A1"].font = heading_font(color_hex=primary_hex, size=18)
        ws.row_dimensions[1].height = 28

        # Column widths.
        widths = {1: 36}
        for col in range(2, 2 + proj_years + 2):  # +2 for terminal columns
            widths[col] = 14
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # Year header row (row 3): FY+1 … FY+N plus a "Terminal" column.
        c = ws.cell(row=3, column=1, value="Period")
        c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
        c.fill = section_fill(); c.border = thin_border()
        c.alignment = left_align()
        for j in range(1, proj_years + 1):
            c = ws.cell(row=3, column=1 + j, value=f"FY+{j}")
            c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
            c.fill = section_fill(); c.border = thin_border()
            c.alignment = right_align()
        ws.row_dimensions[3].height = 22

        cell_count = 0

        # ---- Revenue / EBITDA links (rows 5-6) ----
        style_label(ws.cell(row=5, column=1), value="Revenue")
        style_label(ws.cell(row=6, column=1), value="EBITDA")
        for j in range(1, proj_years + 1):
            col = 1 + j
            col_letter = get_column_letter(col)
            rev_ref = (
                cell_registry.get(f"revenue_y{j}") if cell_registry is not None else None
            )
            ebitda_ref = (
                cell_registry.get(f"ebitda_y{j}") if cell_registry is not None else None
            )
            rev_cell = ws.cell(row=5, column=col)
            eb_cell = ws.cell(row=6, column=col)
            for cell, ref in ((rev_cell, rev_ref), (eb_cell, ebitda_ref)):
                if ref:
                    cell.value = f"={ref}"
                    cell.font = link_font()
                else:
                    cell.value = "(unavailable)"
                    cell.font = muted_font()
                cell.alignment = right_align()
                cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
                cell.border = thin_border()
        cell_count += 2 * proj_years

        # ---- Tax row (= EBITDA × Assumptions!tax_rate) ----
        tax_row = 7
        style_label(ws.cell(row=tax_row, column=1), value="Less: Tax")
        tax_ref = (
            cell_registry.get("tax_rate") if cell_registry is not None
            else absolute_ref("Assumptions", "B16")
        )
        for j in range(1, proj_years + 1):
            col = 1 + j
            col_letter = get_column_letter(col)
            cell = ws.cell(row=tax_row, column=col)
            cell.value = f"={col_letter}6*{tax_ref}"
            cell.font = formula_font()
            cell.alignment = right_align()
            cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
            cell.border = thin_border()

        # ---- D&A row (input % × revenue) ----
        dna_pct_row = 8
        style_label(ws.cell(row=dna_pct_row, column=1), value="Plus: D&A (% of revenue)")
        for j in range(1, proj_years + 1):
            col = 1 + j
            cell = ws.cell(row=dna_pct_row, column=col)
            style_input_cell(cell, value=_DEFAULT_DNA_PCT, number_format=NUMBER_FORMATS["percent"])

        # ---- Capex row (input % × revenue) ----
        capex_pct_row = 9
        capex_default = _parse_capex_intensity(payload)
        style_label(ws.cell(row=capex_pct_row, column=1), value="Less: Capex (% of revenue)")
        for j in range(1, proj_years + 1):
            col = 1 + j
            cell = ws.cell(row=capex_pct_row, column=col)
            style_input_cell(cell, value=capex_default, number_format=NUMBER_FORMATS["percent"])

        # ---- ΔNWC row (link to Working Capital sheet) ----
        nwc_row = 10
        style_label(ws.cell(row=nwc_row, column=1), value="Less: ΔNWC")
        for j in range(1, proj_years + 1):
            col = 1 + j
            cell = ws.cell(row=nwc_row, column=col)
            nwc_ref = (
                cell_registry.get(f"nwc_change_y{j}") if cell_registry is not None
                else None
            )
            if nwc_ref:
                cell.value = f"={nwc_ref}"
                cell.font = link_font()
            else:
                cell.value = 0.0
                cell.font = muted_font()
            cell.alignment = right_align()
            cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
            cell.border = thin_border()

        # ---- Free Cash Flow (= EBITDA - Tax + D&A*Rev - Capex*Rev - ΔNWC) ----
        fcf_row = 11
        style_label(ws.cell(row=fcf_row, column=1), value="Free Cash Flow", bold=True)
        for j in range(1, proj_years + 1):
            col = 1 + j
            col_letter = get_column_letter(col)
            cell = ws.cell(row=fcf_row, column=col)
            cell.value = (
                f"={col_letter}6-{col_letter}{tax_row}"
                f"+{col_letter}5*{col_letter}{dna_pct_row}"
                f"-{col_letter}5*{col_letter}{capex_pct_row}"
                f"-{col_letter}{nwc_row}"
            )
            cell.font = formula_font(bold=True)
            cell.alignment = right_align()
            cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
            cell.border = thin_border()
            if cell_registry is not None:
                cell_registry.set(f"fcf_y{j}", "DCF", f"{col_letter}{fcf_row}")
        cell_count += proj_years

        # ---- Discount factor (= 1 / (1+WACC)^year) ----
        df_row = 13
        style_label(ws.cell(row=df_row, column=1), value="Discount factor")
        wacc_ref = (
            cell_registry.get("wacc") if cell_registry is not None
            else absolute_ref("Assumptions", "B14")
        )
        for j in range(1, proj_years + 1):
            col = 1 + j
            cell = ws.cell(row=df_row, column=col)
            cell.value = f"=1/(1+{wacc_ref})^{j}"
            cell.font = formula_font()
            cell.alignment = right_align()
            cell.number_format = "0.0000"
            cell.border = thin_border()

        # ---- PV of FCF (= FCF × discount_factor) ----
        pv_row = 14
        style_label(ws.cell(row=pv_row, column=1), value="PV of FCF", bold=True)
        for j in range(1, proj_years + 1):
            col = 1 + j
            col_letter = get_column_letter(col)
            cell = ws.cell(row=pv_row, column=col)
            cell.value = f"={col_letter}{fcf_row}*{col_letter}{df_row}"
            cell.font = formula_font(bold=True)
            cell.alignment = right_align()
            cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
            cell.border = thin_border()

        # ---- Section: terminal value ----
        for col in range(1, 2 + proj_years):
            cell_hdr = ws.cell(row=16, column=col)
            cell_hdr.fill = section_fill()
            cell_hdr.border = thin_border()
        ws.cell(row=16, column=1).value = "Terminal value"
        ws.cell(row=16, column=1).font = heading_font(
            color_hex=primary_hex, size=12,
        )

        last_col = 1 + proj_years
        last_col_letter = get_column_letter(last_col)
        tg_ref = (
            cell_registry.get("terminal_growth") if cell_registry is not None
            else absolute_ref("Assumptions", "B15")
        )

        # Gordon Growth: TV = FCF_terminal × (1+g) / (WACC - g)
        # then discounted back by 1/(1+WACC)^N
        gg_row = 17
        style_label(ws.cell(row=gg_row, column=1), value="TV — Gordon Growth (discounted)")
        gg_cell = ws.cell(row=gg_row, column=last_col)
        gg_cell.value = (
            f"={last_col_letter}{fcf_row}*(1+{tg_ref})/({wacc_ref}-{tg_ref})"
            f"*{last_col_letter}{df_row}"
        )
        gg_cell.font = formula_font(bold=True)
        gg_cell.alignment = right_align()
        gg_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
        gg_cell.border = thin_border()

        # Exit Multiple: TV = EBITDA_terminal × exit_multiple
        em_row = 18
        em_input_col = last_col  # default exit multiple lands here
        # Actually use a dedicated input cell for exit multiple.
        em_label_row = 18
        style_label(ws.cell(row=em_label_row, column=1), value="Exit multiple (× EBITDA)")
        ex_cell = ws.cell(row=em_label_row, column=2)
        style_input_cell(ex_cell, value=_DEFAULT_EXIT_MULTIPLE,
                         number_format=NUMBER_FORMATS["multiple"])
        if cell_registry is not None:
            cell_registry.set("exit_multiple", "DCF", "B18")
        em_row = 19
        style_label(ws.cell(row=em_row, column=1), value="TV — Exit Multiple (discounted)")
        em_cell = ws.cell(row=em_row, column=last_col)
        em_cell.value = (
            f"={last_col_letter}6*$B$18*{last_col_letter}{df_row}"
        )
        em_cell.font = formula_font(bold=True)
        em_cell.alignment = right_align()
        em_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
        em_cell.border = thin_border()

        # Selected TV (average of the two methods).
        sel_row = 20
        style_label(ws.cell(row=sel_row, column=1), value="TV — selected (average)")
        sel_cell = ws.cell(row=sel_row, column=last_col)
        sel_cell.value = f"=AVERAGE({last_col_letter}{gg_row},{last_col_letter}{em_row})"
        sel_cell.font = formula_font(bold=True)
        sel_cell.alignment = right_align()
        sel_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
        sel_cell.border = thin_border()

        # ---- Enterprise Value / Equity Value (rows 22-24) ----
        for col in range(1, 2 + proj_years):
            cell_hdr = ws.cell(row=22, column=col)
            cell_hdr.fill = section_fill()
            cell_hdr.border = thin_border()
        ws.cell(row=22, column=1).value = "Valuation"
        ws.cell(row=22, column=1).font = heading_font(
            color_hex=primary_hex, size=12,
        )

        # Enterprise Value (SUM of PV of FCFs + selected TV)
        ev_row = 23
        style_label(ws.cell(row=ev_row, column=1), value="Enterprise Value", bold=True)
        first_pv = get_column_letter(2)
        ev_cell = ws.cell(row=ev_row, column=2)
        ev_cell.value = (
            f"=SUM({first_pv}{pv_row}:{last_col_letter}{pv_row})+{last_col_letter}{sel_row}"
        )
        ev_cell.font = formula_font(bold=True)
        ev_cell.alignment = right_align()
        ev_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
        ev_cell.border = thin_border()
        if cell_registry is not None:
            cell_registry.set("enterprise_value", "DCF", "B23")

        # Less: Net Debt (BLUE assumption — payload.debt_structure is
        # narrative-only; spec hard rule).
        nd_row = 24
        style_label(ws.cell(row=nd_row, column=1), value="Less: Net Debt")
        nd_cell = ws.cell(row=nd_row, column=2)
        style_input_cell(nd_cell, value=_DEFAULT_NET_DEBT,
                         number_format=NUMBER_FORMATS["currency_gbp_m"])
        # Note in column D pointing at the narrative.
        fp = payload_get(payload, "financial_profile", default=None) or {}
        debt_blurb = str(fp.get("debt_structure") or "").strip()
        note_cell = ws.cell(row=nd_row, column=4)
        note_cell.value = (
            "ASSUMPTION — payload.debt_structure is freeform; consultant input required."
            + (f"  Source narrative: \"{debt_blurb[:120]}\"" if debt_blurb else "")
        )
        note_cell.font = muted_font(size=10)
        note_cell.alignment = left_align()

        # Equity Value (= EV - Net Debt)
        eq_row = 25
        style_label(ws.cell(row=eq_row, column=1), value="Equity Value", bold=True)
        eq_cell = ws.cell(row=eq_row, column=2)
        eq_cell.value = f"=B{ev_row}-B{nd_row}"
        eq_cell.font = formula_font(bold=True)
        eq_cell.alignment = right_align()
        eq_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
        eq_cell.border = thin_border()
        if cell_registry is not None:
            cell_registry.set("equity_value", "DCF", f"B{eq_row}")
            cell_registry.set("dna_pct_row", "DCF", str(dna_pct_row))
            cell_registry.set("capex_pct_row", "DCF", str(capex_pct_row))

        return SheetResult(
            sheet_index=workbook.worksheets.index(ws),
            citation_ids=[],
            cell_count=cell_count,
        )
