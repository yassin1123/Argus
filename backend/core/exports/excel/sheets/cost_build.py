"""Cost Build sheet — W12/D2.

Margin profile + opex assumptions + EBITDA / FCF projection.

Structure:
  Row 1:        Title banner.
  Row 3:        Year header row matching Revenue Build (FY+0, FY+1, …).
  Rows 5..N:    Historical margin profile from
                payload.financial_profile.margin_profile (gross margin
                %, EBITDA margin %, FCF margin %). Each row carries
                payload-derived historical values + citation comments
                on the latest-period cell.
  Rows N+1..M:  Projection block — for each projection year Y+1..Y+N:
                  Revenue (link to Revenue Build!revenue_yN)
                  EBITDA margin (link to Assumptions!ebitda_margin_yN)
                  EBITDA = Revenue × Margin (formula)
                  Implied opex = Revenue - EBITDA (formula)

Cross-sheet refs come from the WorkbookBuilder's CellRegistry. If
Revenue Build registered revenue_y1..yN entries, Cost Build references
those; otherwise the projection rows fall back to a "(Revenue Build
unavailable)" note so the workbook still ships.

Spec hard rules:
  - No DCF / NPV today (D3 work).
  - No fabricated margin defaults — historicals come from payload,
    forward margins from Assumptions (which were flagged ASSUMPTION
    on the Assumptions sheet).
"""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter

from ..._base import payload_get
from .._refs import NUMBER_FORMATS, absolute_ref, cell_ref
from .._styles import (
    HEADING_TEXT_HEX,
    MUTED_TEXT_HEX,
    formula_font,
    heading_font,
    left_align,
    link_font,
    muted_font,
    right_align,
    section_fill,
    style_label,
    thin_border,
)
from ..citation_helpers import add_citation_comment, breadcrumb_for_citation
from ._base import SheetBuilderBase, SheetResult
from ._registry import register_sheet


_MARGIN_LABELS: dict[str, str] = {
    "gross_margin":  "Gross margin",
    "ebitda_margin": "EBITDA margin",
    "fcf_margin":    "FCF margin",
}


def _projection_years_for(mode: str | None) -> int:
    """Mirror the Revenue Build horizon so the year columns line up."""
    from .revenue_build import _projection_years_for as _resolve
    return _resolve(mode)


def _historical_periods(payload: Any) -> list[str]:
    """Pull the trajectory period labels from
    payload.financial_profile.revenue_trajectory so the column header
    matches Revenue Build."""
    fp = payload_get(payload, "financial_profile", default=None)
    if not isinstance(fp, dict):
        return []
    traj = fp.get("revenue_trajectory") or {}
    if not isinstance(traj, dict):
        return []
    from ...one_pager_renderer import _coerce_to_list

    out: list[str] = []
    for p in _coerce_to_list(traj.get("points") or [])[-5:]:
        if isinstance(p, dict):
            period = str(p.get("period") or "").strip()
            if period:
                out.append(period)
    return out


def _parse_pct(raw: Any) -> float | None:
    """``"36.4%"`` → ``0.364``; returns None on bad input."""
    if raw is None:
        return None
    s = str(raw).strip().rstrip("%")
    if not s:
        return None
    try:
        return float(s) / 100.0
    except ValueError:
        return None


def _margin_profile(payload: Any) -> dict[str, float]:
    """Decimal-form margin profile from
    ``payload.financial_profile.margin_profile``. Returns an empty
    dict when missing — caller surfaces a placeholder row."""
    fp = payload_get(payload, "financial_profile", default=None)
    if not isinstance(fp, dict):
        return {}
    mp = fp.get("margin_profile") or {}
    if not isinstance(mp, dict):
        return {}
    out: dict[str, float] = {}
    for key in _MARGIN_LABELS:
        v = _parse_pct(mp.get(key))
        if v is not None:
            out[key] = v
    return out


def _build_citation_index(citations: list[Any]) -> dict[str, str]:
    out: dict[str, str] = {}
    for c in citations or []:
        cid = (getattr(c, "claim_id", "") or "").strip()
        if cid and cid not in out:
            out[cid] = breadcrumb_for_citation(c)
    return out


def _latest_ebitda_citation(payload: Any) -> str:
    fp = payload_get(payload, "financial_profile", default=None)
    if not isinstance(fp, dict):
        return ""
    traj = fp.get("ebitda_trajectory") or {}
    if not isinstance(traj, dict):
        return ""
    from ...one_pager_renderer import _coerce_to_list

    pts = _coerce_to_list(traj.get("points") or [])
    if not pts:
        return ""
    last = pts[-1]
    if not isinstance(last, dict):
        return ""
    return str(last.get("source_citation") or "").strip()


@register_sheet("cost_build")
class CostBuildSheet(SheetBuilderBase):
    def build(
        self,
        workbook: Any,
        payload: Any,
        firm_branding: dict[str, Any],
        citations: list[Any],
        cell_registry: Any = None,
    ) -> SheetResult:
        ws = workbook.create_sheet("Cost Build")

        # Resolve mode + horizon (matches Revenue Build).
        mode_hint = payload_get(payload, "_mode_hint", default=None)
        explicit_mode = str(payload_get(payload, "mode", default="") or "").strip()
        mode = mode_hint or explicit_mode or "general"
        proj_years = _projection_years_for(mode)

        primary_hex = str(
            (firm_branding or {}).get("primary_color") or f"#{HEADING_TEXT_HEX}"
        ).lstrip("#")
        ws["A1"] = "Cost Build"
        ws["A1"].font = heading_font(color_hex=primary_hex, size=18)
        ws.row_dimensions[1].height = 28

        historicals = _historical_periods(payload)
        n_hist = len(historicals)

        # Column widths.
        widths = {1: 32}
        for col in range(2, 2 + n_hist + proj_years):
            widths[col] = 14
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # Year header row.
        c = ws.cell(row=3, column=1, value="Period")
        c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
        c.fill = section_fill(); c.border = thin_border()
        c.alignment = left_align()
        for i, period in enumerate(historicals):
            c = ws.cell(row=3, column=2 + i, value=period)
            c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
            c.fill = section_fill(); c.border = thin_border()
            c.alignment = right_align()
        for j in range(1, proj_years + 1):
            c = ws.cell(row=3, column=2 + n_hist + j - 1, value=f"FY+{j}")
            c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
            c.fill = section_fill(); c.border = thin_border()
            c.alignment = right_align()
        ws.row_dimensions[3].height = 22

        cited: list[str] = []
        cell_count = 0
        cite_index = _build_citation_index(citations)
        ebitda_cid = _latest_ebitda_citation(payload)

        # ------------------------------------------------------------------
        # SECTION 1 — Historical margin profile (rows 5-7)
        # ------------------------------------------------------------------
        margins = _margin_profile(payload)
        margin_rows: dict[str, int] = {}
        for i, (key, label) in enumerate(_MARGIN_LABELS.items()):
            row = 5 + i
            margin_rows[key] = row
            style_label(ws.cell(row=row, column=1), value=label)
            if key in margins:
                # Historical: same value across history columns (no
                # per-year margin trajectory on the payload yet — the
                # schema only exposes a latest-period margin profile).
                for i_hist in range(n_hist):
                    cell = ws.cell(row=row, column=2 + i_hist)
                    cell.value = margins[key]
                    cell.font = formula_font()
                    cell.alignment = right_align()
                    cell.number_format = NUMBER_FORMATS["percent"]
                    cell.border = thin_border()
                # Citation on the latest period only (most-recent
                # margin profile is what the payload's source backs).
                if ebitda_cid and key == "ebitda_margin" and n_hist > 0:
                    latest_cell = ws.cell(row=row, column=2 + n_hist - 1)
                    breadcrumb = cite_index.get(ebitda_cid) or ebitda_cid
                    add_citation_comment(
                        latest_cell, claim_id=ebitda_cid,
                        citation_text=breadcrumb,
                    )
                    if ebitda_cid not in cited:
                        cited.append(ebitda_cid)
                # Forward-year margin cells link to Assumptions.
                for j in range(1, proj_years + 1):
                    col = 2 + n_hist + j - 1
                    cell = ws.cell(row=row, column=col)
                    if key == "ebitda_margin" and cell_registry is not None:
                        ref = cell_registry.get(f"ebitda_margin_y{j}") \
                            or absolute_ref("Assumptions", f"B{20 + j}")
                        cell.value = f"={ref}"
                        cell.font = link_font()
                    else:
                        # No projection for gross / FCF margins today.
                        cell.value = None
                    cell.alignment = right_align()
                    cell.number_format = NUMBER_FORMATS["percent"]
                    cell.border = thin_border()
                cell_count += n_hist + proj_years
            else:
                ws.cell(row=row, column=2).value = (
                    "n/a — margin_profile not produced for this engagement"
                )
                ws.cell(row=row, column=2).font = muted_font()

        # ------------------------------------------------------------------
        # SECTION 2 — Forward EBITDA projection (rows 9-11)
        # ------------------------------------------------------------------
        # Row 9: Revenue (link to Revenue Build)
        # Row 10: EBITDA margin (link to historical/projection row above)
        # Row 11: EBITDA = Revenue × EBITDA margin
        # Row 12: Implied opex = Revenue - EBITDA
        rev_row = 9
        em_row = 10
        ebitda_row = 11
        opex_row = 12

        # Section header.
        for col in (1, 2, 3, 4):
            cell_hdr = ws.cell(row=rev_row - 1, column=col)
            cell_hdr.fill = section_fill()
            cell_hdr.border = thin_border()
        ws.cell(row=rev_row - 1, column=1).value = "EBITDA projection"
        ws.cell(row=rev_row - 1, column=1).font = heading_font(
            color_hex=primary_hex, size=12,
        )

        style_label(ws.cell(row=rev_row, column=1), value="Revenue")
        style_label(ws.cell(row=em_row, column=1), value="EBITDA margin")
        style_label(ws.cell(row=ebitda_row, column=1), value="EBITDA")
        style_label(ws.cell(row=opex_row, column=1), value="Implied opex")

        for j in range(1, proj_years + 1):
            col = 2 + n_hist + j - 1
            col_letter = get_column_letter(col)

            # Revenue link
            rev_ref = (
                cell_registry.get(f"revenue_y{j}") if cell_registry is not None else None
            )
            rev_cell = ws.cell(row=rev_row, column=col)
            if rev_ref:
                rev_cell.value = f"={rev_ref}"
                rev_cell.font = link_font()
            else:
                rev_cell.value = "(Revenue Build unavailable)"
                rev_cell.font = muted_font()
            rev_cell.alignment = right_align()
            rev_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
            rev_cell.border = thin_border()

            # EBITDA margin link (forward years from Assumptions).
            em_ref = (
                cell_registry.get(f"ebitda_margin_y{j}") if cell_registry is not None
                else absolute_ref("Assumptions", f"B{20 + j}")
            )
            em_cell = ws.cell(row=em_row, column=col)
            em_cell.value = f"={em_ref}"
            em_cell.font = link_font()
            em_cell.alignment = right_align()
            em_cell.number_format = NUMBER_FORMATS["percent"]
            em_cell.border = thin_border()

            # EBITDA = Revenue * EBITDA margin
            eb_cell = ws.cell(row=ebitda_row, column=col)
            eb_cell.value = (
                f"={col_letter}{rev_row}*{col_letter}{em_row}"
            )
            eb_cell.font = formula_font(bold=True)
            eb_cell.alignment = right_align()
            eb_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
            eb_cell.border = thin_border()

            # Implied opex = Revenue - EBITDA
            opex_cell = ws.cell(row=opex_row, column=col)
            opex_cell.value = (
                f"={col_letter}{rev_row}-{col_letter}{ebitda_row}"
            )
            opex_cell.font = formula_font()
            opex_cell.alignment = right_align()
            opex_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
            opex_cell.border = thin_border()

            cell_count += 4

            # Register EBITDA cell for the W12/D3 DCF sheet.
            if cell_registry is not None:
                cell_registry.set(
                    f"ebitda_y{j}", "Cost Build", f"{col_letter}{ebitda_row}",
                )

        return SheetResult(
            sheet_index=workbook.worksheets.index(ws),
            citation_ids=cited,
            cell_count=cell_count,
        )
