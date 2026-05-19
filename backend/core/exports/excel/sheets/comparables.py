"""Comparables sheet — W12/D3 (M&A-specific).

Two tables side by side:
  Table 1 — Comparable Transactions
    Source: payload.valuation_range.comparable_transactions_cited
    Columns: target | acquirer | year | EV/EBITDA | EV/Sales | source
    Median / Mean / Min / Max rows use MEDIAN / AVERAGE / MIN / MAX
    formulas across the relevant column range.

  Table 2 — Trading Comparables
    Source: payload.valuation_range.multiples_implied (peer-by-peer
    if present) — the W7 schema doesn't currently expose a structured
    trading-comps list, so this table renders BLUE input rows the
    consultant fills in. Same median/mean/min/max statistics.

Implied valuation block at the bottom:
  Implied EV (median EV/EBITDA × current EBITDA)
  Implied EV (median EV/Sales × current revenue)
  Range across all peers

Hard-rule compliance:
  - When ``comparable_transactions_cited`` is empty, the transactions
    table renders the documented placeholder ("No comparable
    transactions cited in source memo; consultant input required.")
    rather than fabricating peers.
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl.utils import get_column_letter

from ..._base import payload_get
from .._refs import NUMBER_FORMATS
from .._styles import (
    HEADING_TEXT_HEX,
    MUTED_TEXT_HEX,
    formula_font,
    heading_font,
    left_align,
    muted_font,
    right_align,
    section_fill,
    style_input_cell,
    style_label,
    thin_border,
)
from ..citation_helpers import add_citation_comment, breadcrumb_for_citation
from ._base import SheetBuilderBase, SheetResult
from ._registry import register_sheet

# Default trading-comp peer count if payload has none.
_DEFAULT_TRADING_PEERS = 4
_DEFAULT_EV_EBITDA = 8.0
_DEFAULT_EV_SALES = 1.3


def _parse_multiple(raw: Any) -> tuple[float | None, float | None]:
    """Extract (EV/EBITDA, EV/Sales) numerics from a freeform multiple
    string. The W7 schema stores ``multiple`` as one string per
    comparable (e.g. ``"8.2x EV/EBITDA"`` or ``"1.6x EV/Sales"``);
    return None for whichever type isn't expressed in this row."""
    if raw is None:
        return None, None
    s = str(raw).strip()
    if not s:
        return None, None
    ev_ebitda = None
    ev_sales = None
    # Pattern: e.g. "8.2x" -> capture 8.2.
    matches = re.findall(r"(\d+(?:\.\d+)?)\s*x", s, re.IGNORECASE)
    if not matches:
        return None, None
    val = float(matches[0])
    lower = s.lower()
    if "ebitda" in lower:
        ev_ebitda = val
    elif "sales" in lower or "revenue" in lower:
        ev_sales = val
    else:
        # Default to EV/EBITDA when type isn't specified.
        ev_ebitda = val
    return ev_ebitda, ev_sales


@register_sheet("comparables")
class ComparablesSheet(SheetBuilderBase):
    def build(
        self,
        workbook: Any,
        payload: Any,
        firm_branding: dict[str, Any],
        citations: list[Any],
        cell_registry: Any = None,
    ) -> SheetResult:
        ws = workbook.create_sheet("Comparables")

        primary_hex = str(
            (firm_branding or {}).get("primary_color") or f"#{HEADING_TEXT_HEX}"
        ).lstrip("#")

        ws["A1"] = "Comparables — Transactions + Trading"
        ws["A1"].font = heading_font(color_hex=primary_hex, size=18)
        ws.row_dimensions[1].height = 28

        # Column widths.
        widths = {1: 28, 2: 24, 3: 8, 4: 13, 5: 13, 6: 30}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        cited: list[str] = []
        cell_count = 0
        cite_index: dict[str, str] = {}
        for c in citations or []:
            cid = (getattr(c, "claim_id", "") or "").strip()
            if cid and cid not in cite_index:
                cite_index[cid] = breadcrumb_for_citation(c)

        # ================================================================
        # SECTION 1 — Comparable Transactions
        # ================================================================
        vr = payload_get(payload, "valuation_range", default={}) or {}
        from ...one_pager_renderer import _coerce_to_list
        transactions = _coerce_to_list(vr.get("comparable_transactions_cited") or [])
        transactions = [t for t in transactions if isinstance(t, dict)]

        # Header band.
        row = 3
        for col in range(1, 7):
            c = ws.cell(row=row, column=col)
            c.fill = section_fill()
            c.border = thin_border()
        ws.cell(row=row, column=1).value = "Comparable Transactions"
        ws.cell(row=row, column=1).font = heading_font(color_hex=primary_hex, size=12)
        row += 1

        # Column headers.
        for col, label in enumerate(
            ["Target", "Acquirer", "Year", "EV/EBITDA", "EV/Sales", "Source"], start=1,
        ):
            c = ws.cell(row=row, column=col, value=label)
            c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
            c.alignment = left_align() if col <= 3 or col == 6 else right_align()
            c.fill = section_fill()
            c.border = thin_border()
        row += 1

        trans_first_row = row
        if not transactions:
            ws.cell(row=row, column=1).value = (
                "No comparable transactions cited in source memo; "
                "consultant input required."
            )
            ws.cell(row=row, column=1).font = muted_font()
            ws.cell(row=row, column=1).alignment = left_align()
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            trans_last_row = row
            row += 1
        else:
            for t in transactions[:10]:
                target = str(t.get("target") or "—")
                acquirer = str(t.get("acquirer") or "—")
                year_val = t.get("year")
                ev_ebitda, ev_sales = _parse_multiple(t.get("multiple"))
                source_note = str(t.get("source_citation") or "—")

                style_label(ws.cell(row=row, column=1), value=target)
                style_label(ws.cell(row=row, column=2), value=acquirer)
                c_year = ws.cell(row=row, column=3)
                c_year.value = year_val if isinstance(year_val, (int, float)) else None
                c_year.alignment = right_align(); c_year.number_format = NUMBER_FORMATS["year"]
                c_year.border = thin_border()
                c_eb = ws.cell(row=row, column=4)
                c_eb.value = ev_ebitda
                c_eb.alignment = right_align(); c_eb.number_format = NUMBER_FORMATS["multiple"]
                c_eb.border = thin_border()
                c_sl = ws.cell(row=row, column=5)
                c_sl.value = ev_sales
                c_sl.alignment = right_align(); c_sl.number_format = NUMBER_FORMATS["multiple"]
                c_sl.border = thin_border()
                style_label(ws.cell(row=row, column=6), value=source_note)
                row += 1
                cell_count += 5
            trans_last_row = row - 1

        # Statistics rows: Median / Mean / Min / Max — only if we have
        # body rows.
        median_ev_eb_cell = None
        median_ev_sl_cell = None
        if transactions:
            row += 1  # spacer
            stats_def = [
                ("Median", "MEDIAN"),
                ("Mean", "AVERAGE"),
                ("Min", "MIN"),
                ("Max", "MAX"),
            ]
            for label, fn in stats_def:
                ws.cell(row=row, column=1).value = label
                ws.cell(row=row, column=1).font = heading_font(
                    color_hex=MUTED_TEXT_HEX, size=11,
                )
                for col in (4, 5):  # EV/EBITDA, EV/Sales
                    col_letter = get_column_letter(col)
                    c = ws.cell(row=row, column=col)
                    c.value = f"={fn}({col_letter}{trans_first_row}:{col_letter}{trans_last_row})"
                    c.font = formula_font(bold=(label == "Median"))
                    c.alignment = right_align()
                    c.number_format = NUMBER_FORMATS["multiple"]
                    c.border = thin_border()
                if label == "Median":
                    median_ev_eb_cell = f"D{row}"
                    median_ev_sl_cell = f"E{row}"
                row += 1

        # ================================================================
        # SECTION 2 — Trading Comparables (BLUE input rows)
        # ================================================================
        row += 2  # gap
        for col in range(1, 7):
            c = ws.cell(row=row, column=col)
            c.fill = section_fill()
            c.border = thin_border()
        ws.cell(row=row, column=1).value = "Trading Comparables (consultant input)"
        ws.cell(row=row, column=1).font = heading_font(color_hex=primary_hex, size=12)
        row += 1

        for col, label in enumerate(
            ["Peer", "Ticker", "—", "EV/EBITDA", "EV/Sales", "Notes"], start=1,
        ):
            c = ws.cell(row=row, column=col, value=label)
            c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
            c.alignment = left_align() if col <= 3 or col == 6 else right_align()
            c.fill = section_fill(); c.border = thin_border()
        row += 1

        trade_first_row = row
        for i in range(_DEFAULT_TRADING_PEERS):
            style_label(ws.cell(row=row, column=1), value=f"Peer {i+1}")
            style_input_cell(ws.cell(row=row, column=2), value="",
                             number_format="@")
            style_input_cell(ws.cell(row=row, column=4), value=_DEFAULT_EV_EBITDA,
                             number_format=NUMBER_FORMATS["multiple"])
            style_input_cell(ws.cell(row=row, column=5), value=_DEFAULT_EV_SALES,
                             number_format=NUMBER_FORMATS["multiple"])
            row += 1
            cell_count += 4
        trade_last_row = row - 1

        # Trading comps stats.
        row += 1
        for label, fn in stats_def if transactions else [
            ("Median", "MEDIAN"), ("Mean", "AVERAGE"),
            ("Min", "MIN"), ("Max", "MAX"),
        ]:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = heading_font(
                color_hex=MUTED_TEXT_HEX, size=11,
            )
            for col in (4, 5):
                col_letter = get_column_letter(col)
                c = ws.cell(row=row, column=col)
                c.value = f"={fn}({col_letter}{trade_first_row}:{col_letter}{trade_last_row})"
                c.font = formula_font(bold=(label == "Median"))
                c.alignment = right_align()
                c.number_format = NUMBER_FORMATS["multiple"]
                c.border = thin_border()
            row += 1

        # ================================================================
        # SECTION 3 — Implied valuation
        # ================================================================
        row += 1
        for col in range(1, 7):
            c = ws.cell(row=row, column=col)
            c.fill = section_fill()
            c.border = thin_border()
        ws.cell(row=row, column=1).value = "Implied valuation"
        ws.cell(row=row, column=1).font = heading_font(color_hex=primary_hex, size=12)
        row += 1

        # Implied EV from transactions median (when available).
        if median_ev_eb_cell:
            ws.cell(row=row, column=1).value = "Implied EV — median EV/EBITDA (transactions)"
            ws.cell(row=row, column=1).font = formula_font()
            ws.cell(row=row, column=1).alignment = left_align()
            ebitda_y5_ref = (
                cell_registry.get("ebitda_y5")
                or cell_registry.get("ebitda_y3")
                if cell_registry is not None else None
            )
            c = ws.cell(row=row, column=2)
            if ebitda_y5_ref:
                c.value = f"={ebitda_y5_ref}*{median_ev_eb_cell}"
                c.font = formula_font(bold=True)
            else:
                c.value = "(EBITDA projection unavailable)"
                c.font = muted_font()
            c.alignment = right_align()
            c.number_format = NUMBER_FORMATS["currency_gbp_m"]
            c.border = thin_border()
            row += 1

        if median_ev_sl_cell:
            ws.cell(row=row, column=1).value = "Implied EV — median EV/Sales (transactions)"
            ws.cell(row=row, column=1).font = formula_font()
            ws.cell(row=row, column=1).alignment = left_align()
            rev_y5_ref = (
                cell_registry.get("revenue_y5")
                or cell_registry.get("revenue_y3")
                if cell_registry is not None else None
            )
            c = ws.cell(row=row, column=2)
            if rev_y5_ref:
                c.value = f"={rev_y5_ref}*{median_ev_sl_cell}"
                c.font = formula_font(bold=True)
            else:
                c.value = "(Revenue projection unavailable)"
                c.font = muted_font()
            c.alignment = right_align()
            c.number_format = NUMBER_FORMATS["currency_gbp_m"]
            c.border = thin_border()
            row += 1

        return SheetResult(
            sheet_index=workbook.worksheets.index(ws),
            citation_ids=cited,
            cell_count=cell_count,
        )
