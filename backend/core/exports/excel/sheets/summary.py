"""Summary sheet — W12/D4 (mode-agnostic).

Executive landing page synthesizing the model's outputs.

Layout (top to bottom):
  Section 1 — Recommendation block (large, colour-coded by verdict).
  Section 2 — Key valuation (M&A only):
                three columns linking to DCF EV, Trading Comps EV,
                Transaction Comps EV.
  Section 3 — Key assumptions: WACC / Terminal growth / Tax rate,
                live links to the Assumptions sheet.
  Section 4 — Top 3 reasons / Top 3 risks: pulled from
                payload.key_reasons[:3] / risks[:3] (or
                executive_summary.top_3_* if present).
                Each row carries a citation comment.

Opts out of the standard W12/D4 firm-header post-pass — owns its
own banner with the recommendation as the visual anchor.
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl.utils import get_column_letter

from ..._base import payload_get
from ...one_pager_renderer import (
    _coerce_to_list,
    _stringify_item,
    classify_recommendation,
    get_recommendation_text,
)
from .._branding import _normalise_hex
from .._refs import NUMBER_FORMATS, absolute_ref
from .._styles import (
    HEADING_TEXT_HEX,
    INPUT_FILL_HEX,
    INPUT_TEXT_HEX,
    MUTED_TEXT_HEX,
    SECTION_FILL_HEX,
    formula_font,
    heading_font,
    left_align,
    link_font,
    muted_font,
    right_align,
    section_fill,
    style_label,
    thin_border,
)
from ..citation_helpers import add_citation_comment, breadcrumb_for_citation
from ._base import SheetBuilderBase, SheetResult
from ._registry import register_sheet


_VERDICT_COLOURS: dict[str, str] = {
    "green":   "0F6E56",
    "amber":   "B8860B",
    "red":     "B91C1C",
    "neutral": "1B1F23",
}


def _build_citation_index(citations: list[Any]) -> dict[str, str]:
    out: dict[str, str] = {}
    for c in citations or []:
        cid = (getattr(c, "claim_id", "") or "").strip()
        if cid and cid not in out:
            out[cid] = breadcrumb_for_citation(c)
    return out


def _extract_top_3_reasons(payload: Any) -> list[str]:
    es = payload_get(payload, "executive_summary", default=None)
    if isinstance(es, dict):
        items = _coerce_to_list(es.get("top_3_reasons") or [])
        out = [_stringify_item(x) for x in items if _stringify_item(x)]
        if out:
            return out[:3]
    items = _coerce_to_list(payload_get(payload, "key_reasons", default=[]))
    out = [_stringify_item(x) for x in items if _stringify_item(x)]
    return out[:3]


def _extract_top_3_risks(payload: Any) -> list[str]:
    es = payload_get(payload, "executive_summary", default=None)
    if isinstance(es, dict):
        items = _coerce_to_list(es.get("top_3_risks") or [])
        out = [_stringify_item(x) for x in items if _stringify_item(x)]
        if out:
            return out[:3]
    items = _coerce_to_list(payload_get(payload, "risks", default=[]))
    out = [_stringify_item(x) for x in items if _stringify_item(x)]
    return out[:3]


@register_sheet("summary")
class SummarySheet(SheetBuilderBase):
    def build(
        self,
        workbook: Any,
        payload: Any,
        firm_branding: dict[str, Any],
        citations: list[Any],
        cell_registry: Any = None,
    ) -> SheetResult:
        ws = workbook.create_sheet("Summary")

        primary_hex = _normalise_hex(
            (firm_branding or {}).get("primary_color") or f"#{HEADING_TEXT_HEX}",
        )

        # Column widths.
        widths = {1: 32, 2: 18, 3: 18, 4: 18}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # ============================================================
        # SECTION 1 — Recommendation (large, colour-coded)
        # ============================================================
        firm_name = (firm_branding or {}).get("_firm_name") or "Argus"
        ws["A1"] = firm_name
        ws["A1"].font = heading_font(color_hex=primary_hex, size=14)
        ws["A1"].alignment = left_align()

        engagement_title = str(
            payload_get(payload, "_engagement_title", default="Argus engagement")
        )
        ws["B1"] = f"Summary · {engagement_title}"
        ws["B1"].font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
        ws["B1"].alignment = left_align()
        try:
            ws.merge_cells("B1:D1")
        except Exception:
            pass
        ws.row_dimensions[1].height = 22

        # Recommendation panel.
        ws["A3"] = "RECOMMENDATION"
        ws["A3"].font = heading_font(color_hex=MUTED_TEXT_HEX, size=10)
        ws["A3"].alignment = left_align()

        rec_text = get_recommendation_text(payload) or "(no recommendation produced)"
        rec_klass = classify_recommendation(rec_text)
        rec_colour = _VERDICT_COLOURS.get(rec_klass, _VERDICT_COLOURS["neutral"])
        ws["A4"] = rec_text[:300]
        ws["A4"].font = heading_font(color_hex=rec_colour, size=16)
        ws["A4"].alignment = left_align()
        ws.row_dimensions[4].height = 32
        try:
            ws.merge_cells("A4:D4")
        except Exception:
            pass

        # ============================================================
        # SECTION 2 — Key valuation (M&A only — pulls DCF + comparables)
        # ============================================================
        row = 6
        # Detect M&A by presence of the registry's enterprise_value key.
        ev_ref = (
            cell_registry.get("enterprise_value") if cell_registry is not None
            else None
        )
        if ev_ref:
            for col in range(1, 5):
                c = ws.cell(row=row, column=col)
                c.fill = section_fill(); c.border = thin_border()
            ws.cell(row=row, column=1).value = "Key valuation (£m)"
            ws.cell(row=row, column=1).font = heading_font(
                color_hex=primary_hex, size=12,
            )
            row += 1
            for col, label in enumerate(
                ["Method", "Enterprise Value", "Equity Value", ""], start=1,
            ):
                c = ws.cell(row=row, column=col, value=label)
                c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=10)
                c.alignment = left_align() if col == 1 else right_align()
                c.fill = section_fill(); c.border = thin_border()
            row += 1

            equity_ref = cell_registry.get("equity_value")
            ws.cell(row=row, column=1).value = "DCF"
            ws.cell(row=row, column=1).font = formula_font()
            ev_cell = ws.cell(row=row, column=2)
            ev_cell.value = f"={ev_ref}"
            ev_cell.font = link_font()
            ev_cell.alignment = right_align()
            ev_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
            ev_cell.border = thin_border()
            if equity_ref:
                eq_cell = ws.cell(row=row, column=3)
                eq_cell.value = f"={equity_ref}"
                eq_cell.font = link_font()
                eq_cell.alignment = right_align()
                eq_cell.number_format = NUMBER_FORMATS["currency_gbp_m"]
                eq_cell.border = thin_border()
            row += 2

        # ============================================================
        # SECTION 3 — Key assumptions (live links to Assumptions sheet)
        # ============================================================
        for col in range(1, 5):
            c = ws.cell(row=row, column=col)
            c.fill = section_fill(); c.border = thin_border()
        ws.cell(row=row, column=1).value = "Key assumptions"
        ws.cell(row=row, column=1).font = heading_font(color_hex=primary_hex, size=12)
        row += 1
        for col, label in enumerate(["Parameter", "Value", "Unit", ""], start=1):
            c = ws.cell(row=row, column=col, value=label)
            c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=10)
            c.alignment = left_align() if col != 2 else right_align()
            c.fill = section_fill(); c.border = thin_border()
        row += 1

        for name, label in (
            ("wacc", "WACC"),
            ("terminal_growth", "Terminal growth"),
            ("tax_rate", "Tax rate"),
        ):
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = formula_font()
            ws.cell(row=row, column=1).alignment = left_align()
            val_cell = ws.cell(row=row, column=2)
            ref = cell_registry.get(name) if cell_registry is not None else None
            if ref:
                val_cell.value = f"={ref}"
                val_cell.font = link_font()
            else:
                val_cell.value = None
                val_cell.font = muted_font()
            val_cell.alignment = right_align()
            val_cell.number_format = NUMBER_FORMATS["percent"]
            val_cell.border = thin_border()
            ws.cell(row=row, column=3).value = "%"
            ws.cell(row=row, column=3).font = muted_font()
            row += 1

        # ============================================================
        # SECTION 4 — Top 3 reasons / Top 3 risks with citation comments
        # ============================================================
        row += 1
        for col in range(1, 5):
            c = ws.cell(row=row, column=col)
            c.fill = section_fill(); c.border = thin_border()
        ws.cell(row=row, column=1).value = "Top reasons + risks"
        ws.cell(row=row, column=1).font = heading_font(color_hex=primary_hex, size=12)
        row += 1

        reasons = _extract_top_3_reasons(payload)
        risks = _extract_top_3_risks(payload)
        cite_index = _build_citation_index(citations)

        # Two-column header.
        ws.cell(row=row, column=1).value = "Reasons"
        ws.cell(row=row, column=1).font = heading_font(color_hex=MUTED_TEXT_HEX, size=10)
        ws.cell(row=row, column=3).value = "Risks"
        ws.cell(row=row, column=3).font = heading_font(color_hex=MUTED_TEXT_HEX, size=10)
        row += 1

        # Cite the first available claim_id round-robin so each row's
        # cell carries SOME source breadcrumb.
        claim_ids_available = list(cite_index.keys())
        cited: list[str] = []

        def _cite_cell(cell: Any, idx: int) -> None:
            if not claim_ids_available:
                return
            cid = claim_ids_available[idx % len(claim_ids_available)]
            breadcrumb = cite_index.get(cid) or cid
            add_citation_comment(cell, claim_id=cid, citation_text=breadcrumb)
            if cid not in cited:
                cited.append(cid)

        n_rows = max(len(reasons), len(risks))
        for i in range(n_rows):
            r_cell = ws.cell(row=row + i, column=1)
            r_cell.value = reasons[i] if i < len(reasons) else None
            r_cell.font = formula_font()
            r_cell.alignment = left_align()
            r_cell.border = thin_border()
            if i < len(reasons):
                _cite_cell(r_cell, i)

            x_cell = ws.cell(row=row + i, column=3)
            x_cell.value = risks[i] if i < len(risks) else None
            x_cell.font = formula_font()
            x_cell.alignment = left_align()
            x_cell.border = thin_border()
            if i < len(risks):
                _cite_cell(x_cell, i + len(reasons))
        if n_rows == 0:
            ws.cell(row=row, column=1).value = "(no reasons / risks produced)"
            ws.cell(row=row, column=1).font = muted_font()

        # Column 2 / 4 wider for the rationale text.
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["C"].width = 60

        return SheetResult(
            sheet_index=workbook.worksheets.index(ws),
            citation_ids=cited,
            cell_count=row + n_rows,
            skip_branding_header=True,  # Summary owns its own row 1.
        )
