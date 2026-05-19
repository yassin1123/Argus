"""Assumptions sheet — W12/D1.

Sheet name: "Assumptions".

Four-column layout:
  A: Parameter name
  B: Value          (input or payload-derived)
  C: Unit
  D: Source / notes (citation comment also attached when payload-derived)

The sheet has two sections:

  1. PAYLOAD-DERIVED VALUES — pulled from
     payload.financial_profile.revenue_trajectory + valuation_range.
     Black-text formula/output styling; cell carries a Comment
     pointing back to the writer's source citation.

  2. CONSULTANT INPUTS — standard DCF assumptions (WACC, terminal
     growth, tax rate, growth projections). Blue-text-on-yellow-fill
     so the user knows "edit these, not the others." Defaults are
     industry-standard placeholders with an "ASSUMPTION — review
     before use" note in column D, per the spec hard rule against
     fabricating values.

Hard rules:
  - Don't fabricate consultant assumptions; defaults are flagged.
  - All payload-derived cells get a citation comment.
"""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter

from ..._base import payload_get
from ...one_pager_renderer import _coerce_to_list
from .._styles import (
    HEADING_TEXT_HEX,
    INPUT_FILL_HEX,
    INPUT_TEXT_HEX,
    MUTED_TEXT_HEX,
    heading_font,
    left_align,
    muted_font,
    right_align,
    section_fill,
    style_input_cell,
    style_label,
    thin_border,
)
from ..citation_helpers import add_citation_comment, breadcrumb_for_citation
from ._base import SheetBuilderBase, SheetResult
from ._registry import register_sheet


# --- Default input assumptions (flagged "ASSUMPTION — review before use") ---
_DEFAULT_WACC = 0.10
_DEFAULT_TERMINAL_GROWTH = 0.025
_DEFAULT_TAX_RATE = 0.25
_DEFAULT_PROJECTED_GROWTH = 0.05  # year-1 projected growth on top of trajectory

_ASSUMPTION_NOTE = "ASSUMPTION — review before use"


def _build_citation_index(citations: list[Any]) -> dict[str, str]:
    """Map claim_id → breadcrumb so payload-derived cells can attach a
    comment without each builder re-reading the citation list."""
    out: dict[str, str] = {}
    for c in citations or []:
        cid = (getattr(c, "claim_id", "") or "").strip()
        if cid and cid not in out:
            out[cid] = breadcrumb_for_citation(c)
    return out


def _latest_trajectory_point(payload: Any, key: str) -> tuple[str, float, list[str]] | None:
    """Return ``(period, value_gbp_m, citation_ids)`` for the latest
    point on the named trajectory (revenue or ebitda). Returns None
    if the trajectory is absent or empty.

    Citation IDs surface from ``point.source_citation`` (and fall
    back to the trajectory's parent block — currently no parent-level
    citation field, so cite-list is single-id at most)."""
    fp = payload_get(payload, "financial_profile", default=None)
    if not isinstance(fp, dict):
        return None
    traj = fp.get(key) or {}
    if not isinstance(traj, dict):
        return None
    pts = _coerce_to_list(traj.get("points") or [])
    if not pts:
        return None
    last = pts[-1]
    if not isinstance(last, dict):
        return None
    try:
        v = float(last.get("value_gbp_m"))
    except (TypeError, ValueError):
        return None
    period = str(last.get("period") or "").strip()
    cid = str(last.get("source_citation") or "").strip()
    return (period or "Latest", v, [cid] if cid else [])


def _default_ebitda_margin_from_payload(payload: Any) -> float:
    """Extract a default EBITDA margin (decimal) from
    ``payload.financial_profile.margin_profile.ebitda_margin`` when
    present. Returns 0.10 when missing — flagged ASSUMPTION on the
    sheet so a consultant knows to override.
    """
    fp = payload_get(payload, "financial_profile", default=None)
    if not isinstance(fp, dict):
        return 0.10
    mp = fp.get("margin_profile") or {}
    if not isinstance(mp, dict):
        return 0.10
    raw = str(mp.get("ebitda_margin") or "").strip().rstrip("%")
    if not raw:
        return 0.10
    try:
        return float(raw) / 100.0
    except ValueError:
        return 0.10


def _valuation_scenario(payload: Any, key: str) -> tuple[float, str] | None:
    """Pull ``(gbp_m, methodology)`` from valuation_range[key]."""
    vr = payload_get(payload, "valuation_range", default=None)
    if not isinstance(vr, dict):
        return None
    node = vr.get(key) or {}
    if not isinstance(node, dict):
        return None
    try:
        v = float(node.get("gbp_m"))
    except (TypeError, ValueError):
        return None
    return v, str(node.get("methodology") or "")


@register_sheet("assumptions")
class AssumptionsSheet(SheetBuilderBase):
    def build(
        self,
        workbook: Any,
        payload: Any,
        firm_branding: dict[str, Any],
        citations: list[Any],
        cell_registry: Any = None,
    ) -> SheetResult:
        ws = workbook.create_sheet("Assumptions")

        # Column widths.
        widths = {1: 40, 2: 18, 3: 12, 4: 60}
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w

        primary_hex = str(
            (firm_branding or {}).get("primary_color") or f"#{HEADING_TEXT_HEX}"
        ).lstrip("#")

        # ---- Header row ----
        ws["A1"] = "Assumptions"
        ws["A1"].font = heading_font(color_hex=primary_hex, size=18)
        ws.row_dimensions[1].height = 28

        # Column headers (row 3).
        headers = ["Parameter", "Value", "Unit", "Source / notes"]
        for col, label in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=label)
            cell.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
            cell.alignment = left_align()
            cell.fill = section_fill()
            cell.border = thin_border()
        ws.row_dimensions[3].height = 22

        cited: list[str] = []
        cell_count = 0
        cite_index = _build_citation_index(citations)

        # =================================================================
        # SECTION 1 — Payload-derived values (black text, citation comments)
        # =================================================================
        row = 5
        _write_section_band(ws, row, "Payload-derived values", primary_hex)
        row += 1

        # Current revenue (latest revenue trajectory point).
        rev = _latest_trajectory_point(payload, "revenue_trajectory")
        if rev is not None:
            period, value, cids = rev
            _write_derived_row(
                ws, row,
                parameter=f"Current revenue ({period})",
                value=value,
                unit="£m",
                source_note="Latest revenue_trajectory point",
                citation_ids=cids,
                cite_index=cite_index,
                cited=cited,
            )
            row += 1; cell_count += 1

        # Current EBITDA.
        ebitda = _latest_trajectory_point(payload, "ebitda_trajectory")
        if ebitda is not None:
            period, value, cids = ebitda
            _write_derived_row(
                ws, row,
                parameter=f"Current EBITDA ({period})",
                value=value,
                unit="£m",
                source_note="Latest ebitda_trajectory point",
                citation_ids=cids,
                cite_index=cite_index,
                cited=cited,
            )
            row += 1; cell_count += 1

        # Valuation range scenarios (low / base / high).
        for scenario in ("low", "base", "high"):
            r = _valuation_scenario(payload, scenario)
            if r is None:
                continue
            value, methodology = r
            _write_derived_row(
                ws, row,
                parameter=f"Valuation — {scenario}",
                value=value,
                unit="£m",
                source_note=methodology or "valuation_range",
                citation_ids=[],
                cite_index=cite_index,
                cited=cited,
            )
            row += 1; cell_count += 1

        # =================================================================
        # SECTION 2 — Consultant inputs (blue text on yellow fill)
        # =================================================================
        row += 1  # spacer
        _write_section_band(ws, row, "Consultant inputs (editable)", primary_hex)
        row += 1

        # Core DCF inputs (used by W12/D3 DCF sheet too).
        for name, label, value in (
            ("wacc",             "WACC",                  _DEFAULT_WACC),
            ("terminal_growth",  "Terminal growth rate",  _DEFAULT_TERMINAL_GROWTH),
            ("tax_rate",         "Tax rate",              _DEFAULT_TAX_RATE),
        ):
            _write_input_row(
                ws, row,
                parameter=label,
                value=value,
                unit="%",
                source_note=_ASSUMPTION_NOTE,
                number_format="0.0%",
            )
            if cell_registry is not None:
                cell_registry.set(name, "Assumptions", f"B{row}")
            row += 1; cell_count += 1

        # Per-year revenue growth (Y+1 … Y+5). Downstream Revenue Build
        # references these per projection column. Flagged
        # ASSUMPTION — spec hard rule "Don't auto-extrapolate growth
        # rates from historical CAGR."
        row += 1  # subsection spacer
        _write_section_band(ws, row, "Revenue growth — projection years", primary_hex)
        row += 1
        for year in range(1, 6):
            _write_input_row(
                ws, row,
                parameter=f"Revenue growth (Y+{year})",
                value=_DEFAULT_PROJECTED_GROWTH,
                unit="%",
                source_note=_ASSUMPTION_NOTE
                + " — single default; consultant tunes per year.",
                number_format="0.0%",
            )
            if cell_registry is not None:
                cell_registry.set(f"revenue_growth_y{year}", "Assumptions", f"B{row}")
            row += 1; cell_count += 1

        # Per-year EBITDA margin (Y+1 … Y+5). Downstream Cost Build
        # references these to project EBITDA off Revenue Build.
        row += 1  # subsection spacer
        _write_section_band(ws, row, "EBITDA margin — projection years", primary_hex)
        row += 1
        _default_ebitda_margin = _default_ebitda_margin_from_payload(payload)
        for year in range(1, 6):
            _write_input_row(
                ws, row,
                parameter=f"EBITDA margin (Y+{year})",
                value=_default_ebitda_margin,
                unit="%",
                source_note=(
                    _ASSUMPTION_NOTE
                    + " — defaulted from payload.financial_profile.margin_profile"
                    + " when available, else 10%."
                ),
                number_format="0.0%",
            )
            if cell_registry is not None:
                cell_registry.set(f"ebitda_margin_y{year}", "Assumptions", f"B{row}")
            row += 1; cell_count += 1

        return SheetResult(
            sheet_index=workbook.worksheets.index(ws),
            citation_ids=cited,
            cell_count=cell_count,
        )


# ----------------------------------------------------------------------------
# Row writers
# ----------------------------------------------------------------------------


def _write_section_band(ws: Any, row: int, title: str, primary_hex: str) -> None:
    """Write a coloured section header band spanning A:D."""
    for col in (1, 2, 3, 4):
        c = ws.cell(row=row, column=col)
        c.fill = section_fill()
        c.border = thin_border()
    ws.cell(row=row, column=1).value = title
    ws.cell(row=row, column=1).font = heading_font(
        color_hex=primary_hex, size=12,
    )
    ws.row_dimensions[row].height = 22


def _write_derived_row(
    ws: Any,
    row: int,
    *,
    parameter: str,
    value: float | int | str,
    unit: str,
    source_note: str,
    citation_ids: list[str],
    cite_index: dict[str, str],
    cited: list[str],
) -> None:
    """Payload-derived row: parameter | value | unit | source_note.

    The value cell uses formula-cell styling (black text on white)
    and carries a Comment with the first citation breadcrumb (if any)
    so a partner can hover and see exactly where the number came from.
    """
    style_label(ws.cell(row=row, column=1), value=parameter)
    val_cell = ws.cell(row=row, column=2)
    val_cell.value = value
    val_cell.alignment = right_align()
    val_cell.border = thin_border()
    if isinstance(value, (int, float)):
        val_cell.number_format = '#,##0.0'

    style_label(ws.cell(row=row, column=3), value=unit)

    note_cell = ws.cell(row=row, column=4)
    note_cell.value = source_note
    note_cell.font = muted_font()
    note_cell.alignment = left_align()
    note_cell.border = thin_border()

    for cid in citation_ids:
        if not cid:
            continue
        breadcrumb = cite_index.get(cid) or cid
        add_citation_comment(val_cell, claim_id=cid, citation_text=breadcrumb)
        if cid not in cited:
            cited.append(cid)


def _write_input_row(
    ws: Any,
    row: int,
    *,
    parameter: str,
    value: float | int | str,
    unit: str,
    source_note: str,
    number_format: str = "General",
) -> None:
    """Consultant input row: parameter | INPUT value | unit | note.

    The value cell uses input-cell styling (blue text on yellow fill)
    so the user immediately sees "this is editable." The note cell
    carries the standard "ASSUMPTION — review before use" string so
    the partner / oncall reviewer knows defaults are placeholders
    rather than verified numbers.
    """
    style_label(ws.cell(row=row, column=1), value=parameter)

    val_cell = ws.cell(row=row, column=2)
    style_input_cell(val_cell, value=value, number_format=number_format)
    # Sanity check for tests: input fill is the canonical yellow.
    # (style_input_cell already sets it; this just documents the contract.)

    style_label(ws.cell(row=row, column=3), value=unit)

    note_cell = ws.cell(row=row, column=4)
    note_cell.value = source_note
    note_cell.font = muted_font()
    note_cell.alignment = left_align()
    note_cell.border = thin_border()
