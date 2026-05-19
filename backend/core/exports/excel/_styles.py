"""Shared Excel-modelling style primitives — W12/D1.

Industry-standard consultant-modelling colour convention:
  - BLUE text on yellow fill        = user-editable INPUT cell
  - BLACK text on white              = formula / output cell
  - GREEN text on white              = cross-sheet link (W12/D3)
  - bold + primary colour on white   = section heading

The names + RGB values here are the single source of truth for the
W12 deck of sheets, and are exported so tests can assert against
them directly (test_assumptions_sheet_has_color_coded_inputs).
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# RGB colour constants (without the leading #).
# Industry-standard Excel modelling palette per the W12 spec:
#   BLUE on YELLOW  = input cells (consultant edits)
#   BLACK on white  = computed cells (formulas)
#   GREEN on white  = link cells (cross-sheet references)
#   GREY            = labels / notes
INPUT_FILL_HEX = "FFFF99"   # industry-standard yellow for inputs
INPUT_TEXT_HEX = "1F4E79"   # blue per W12/D4 spec
FORMULA_TEXT_HEX = "000000"  # black
LINK_TEXT_HEX = "00B050"    # green
HEADING_TEXT_HEX = "0F6E56"  # firm primary (default; overridden per-firm)
MUTED_TEXT_HEX = "5B6470"   # secondary slate for source/notes column
SECTION_FILL_HEX = "F2F4F6"  # very light grey for section bands

DEFAULT_FONT = "Calibri"


def input_font(*, color_hex: str = INPUT_TEXT_HEX, bold: bool = True) -> Font:
    return Font(name=DEFAULT_FONT, size=11, bold=bold, color=color_hex)


def formula_font(*, color_hex: str = FORMULA_TEXT_HEX, bold: bool = False) -> Font:
    return Font(name=DEFAULT_FONT, size=11, bold=bold, color=color_hex)


def link_font() -> Font:
    return Font(name=DEFAULT_FONT, size=11, color=LINK_TEXT_HEX, italic=True)


def heading_font(*, color_hex: str = HEADING_TEXT_HEX, size: int = 14) -> Font:
    return Font(name=DEFAULT_FONT, size=size, bold=True, color=color_hex)


def muted_font(*, size: int = 10) -> Font:
    return Font(name=DEFAULT_FONT, size=size, color=MUTED_TEXT_HEX)


def input_fill() -> PatternFill:
    return PatternFill(start_color=INPUT_FILL_HEX, end_color=INPUT_FILL_HEX,
                       fill_type="solid")


def section_fill() -> PatternFill:
    return PatternFill(start_color=SECTION_FILL_HEX, end_color=SECTION_FILL_HEX,
                       fill_type="solid")


def thin_border(*, bottom: bool = True) -> Border:
    side = Side(border_style="thin", color="D4D8DE")
    return Border(
        top=Side(border_style=None) if not bottom else side,
        bottom=side if bottom else Side(border_style=None),
        left=Side(border_style=None),
        right=Side(border_style=None),
    )


def left_align(*, indent: int = 0) -> Alignment:
    return Alignment(horizontal="left", vertical="center",
                     wrap_text=True, indent=indent)


def right_align() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def style_input_cell(cell: "Cell", *, value, number_format: str = "General",
                     comment: str | None = None) -> None:  # noqa: F821
    """Mark a cell as a user-editable input (blue text on yellow fill)."""
    cell.value = value
    cell.font = input_font()
    cell.fill = input_fill()
    cell.alignment = right_align()
    cell.number_format = number_format
    cell.border = thin_border()


def style_formula_cell(cell: "Cell", *, formula: str,  # noqa: F821
                       number_format: str = "General") -> None:
    """Mark a cell as a derived formula (black, no fill)."""
    cell.value = formula  # openpyxl treats strings starting with '=' as formulas
    cell.font = formula_font()
    cell.alignment = right_align()
    cell.number_format = number_format
    cell.border = thin_border()


def style_heading(cell: "Cell", *, value: str, color_hex: str | None = None,  # noqa: F821
                  size: int = 14) -> None:
    cell.value = value
    cell.font = heading_font(color_hex=color_hex or HEADING_TEXT_HEX, size=size)
    cell.alignment = left_align()


def style_label(cell: "Cell", *, value: str, bold: bool = False) -> None:  # noqa: F821
    cell.value = value
    cell.font = Font(name=DEFAULT_FONT, size=11, bold=bold)
    cell.alignment = left_align()


__all__ = [
    "DEFAULT_FONT",
    "FORMULA_TEXT_HEX",
    "HEADING_TEXT_HEX",
    "INPUT_FILL_HEX",
    "INPUT_TEXT_HEX",
    "LINK_TEXT_HEX",
    "MUTED_TEXT_HEX",
    "SECTION_FILL_HEX",
    "formula_font",
    "heading_font",
    "input_fill",
    "input_font",
    "left_align",
    "link_font",
    "muted_font",
    "right_align",
    "section_fill",
    "style_formula_cell",
    "style_heading",
    "style_input_cell",
    "style_label",
    "thin_border",
]
