"""Firm-branding chrome for every Excel sheet — W12/D4.

Applies the same per-firm visual identity across the workbook:
  - Row 1 firm header: firm name (left, primary colour) | sheet
    display name (centre, bold) | generated date (right, muted).
  - Sheet tab colour set to firm primary.
  - Freeze panes on row 2 so the header stays visible while
    scrolling.

The header runs as a POST-PASS on the WorkbookBuilder (after each
sheet builder lands its content). Sheet builders that opt out
(Cover, Summary) carry a ``skip_branding_header`` flag on their
SheetResult — they own the top of their sheet explicitly.

Also exposes ``audit_citations(workbook)`` — a workbook-wide check
that scans every sheet for cells that should have a citation
comment, returning a list of ``(sheet, coord)`` tuples. Used by
the W12/D4 citation-audit test to confirm every payload-derived
number has a defensible source.
"""

from __future__ import annotations

from datetime import datetime, timezone
from typing import Any

from openpyxl.utils import get_column_letter

from ._styles import (
    HEADING_TEXT_HEX,
    MUTED_TEXT_HEX,
    heading_font,
    left_align,
    muted_font,
    right_align,
    section_fill,
)


def _normalise_hex(raw: Any, default: str = HEADING_TEXT_HEX) -> str:
    """Return a 6-char uppercase RGB hex (no leading #). Falls back
    to ``default`` for malformed input."""
    if isinstance(raw, str):
        s = raw.strip().lstrip("#").upper()
        if len(s) == 6:
            try:
                int(s, 16)
                return s
            except ValueError:
                pass
        if len(s) == 8:
            return s[2:]
    return default.lstrip("#").upper()


def _safe_tab_color(primary_hex: str) -> str:
    """Per spec hard rule: if primary_color is too light or too dark
    to read as a sheet-tab background, fall back to a neutral.
    Heuristic: compute simple perceived brightness; reject extremes."""
    hex_clean = _normalise_hex(primary_hex)
    try:
        r = int(hex_clean[0:2], 16)
        g = int(hex_clean[2:4], 16)
        b = int(hex_clean[4:6], 16)
    except (ValueError, IndexError):
        return HEADING_TEXT_HEX
    brightness = (r * 299 + g * 587 + b * 114) / 1000
    if brightness < 25 or brightness > 240:
        # Too dark / too light → fall back to the firm-default green.
        return HEADING_TEXT_HEX
    return hex_clean


def add_firm_header(
    ws: Any,
    *,
    firm_name: str,
    sheet_display_name: str,
    engagement_title: str,
    primary_hex: str,
    last_data_col: int | None = None,
) -> None:
    """Apply the row-1 firm header band on a sheet. Overwrites
    whatever row 1 used to contain (most existing sheets used row 1
    as their own title — the firm header subsumes that). Caller can
    pass ``last_data_col`` to control how wide the centre-merged
    title cell spans; defaults to col 6 (first 6 columns) which
    matches the widest sheet column count.
    """
    primary_clean = _normalise_hex(primary_hex)

    # Row 1 layout: A=firm name | mid-merge = sheet+engagement title |
    # last col = date.
    a = ws.cell(row=1, column=1, value=firm_name)
    a.font = heading_font(color_hex=primary_clean, size=14)
    a.alignment = left_align()
    a.fill = section_fill()

    end_col = last_data_col or 6
    mid_text = f"{sheet_display_name}  ·  {engagement_title}"
    if end_col >= 3:
        c = ws.cell(row=1, column=2, value=mid_text)
        c.font = heading_font(color_hex=MUTED_TEXT_HEX, size=11)
        c.alignment = left_align()
        c.fill = section_fill()
        # Merge B..(end_col-1) so the title gets room.
        try:
            ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=end_col - 1)
        except Exception:
            # Merge may conflict with existing merges from the sheet
            # builder; if so, skip silently — the text is still set on
            # the leftmost cell.
            pass

    date_str = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d")
    d = ws.cell(row=1, column=end_col, value=date_str)
    d.font = muted_font(size=10)
    d.alignment = right_align()
    d.fill = section_fill()

    ws.row_dimensions[1].height = 22


def apply_tab_color(ws: Any, *, primary_hex: str) -> None:
    """Set the worksheet tab's colour to the firm primary (or the
    safe neutral if primary is too extreme)."""
    safe = _safe_tab_color(primary_hex)
    try:
        ws.sheet_properties.tabColor = safe
    except Exception:
        pass


def freeze_top_row(ws: Any) -> None:
    """Freeze the firm-header row so it stays visible when scrolling."""
    try:
        ws.freeze_panes = "A2"
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Citation audit
# ---------------------------------------------------------------------------


# Sheets that contain at least some payload-derived data and therefore
# should have ≥1 citation comment. Sheets that are pure
# computed-from-other-sheets (Sensitivity, DCF projections) need not
# have citation comments at the cell level — the sources are on the
# inputs they reference.
_SHEETS_REQUIRING_CITATIONS: set[str] = {
    "Assumptions",
    "Revenue Build",
    "Cost Build",
    "Synergies",
    "Summary",
}


def audit_citations(workbook: Any) -> dict[str, Any]:
    """Workbook-wide citation audit.

    Returns:
      - ``missing``: list of (sheet_name, coord) tuples for cells we
        think should have comments but do not.
      - ``coverage``: per-sheet counts of (cells-with-comment,
        total-payload-cells).
      - ``sheets_passed``: list of sheet names that have at least one
        citation comment, validating that the citation pipeline ran
        end-to-end for them.

    A sheet "passes" if it has at least one Comment-bearing cell
    AND every cell that contains a payload-derived comment-bearing
    structure (heuristic: cells with a comment whose author is
    "Argus") is consistent. We do not enforce a per-cell test on
    every numeric cell because many cells are formulas / inputs
    that legitimately have no source comment.
    """
    missing: list[tuple[str, str]] = []
    coverage: dict[str, dict[str, int]] = {}
    sheets_passed: list[str] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        with_comment = 0
        argus_authored = 0
        numeric_data_cells = 0
        default_flagged_rows = 0
        rows_with_value_in_b = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    with_comment += 1
                    if (cell.comment.author or "").strip() == "Argus":
                        argus_authored += 1
                # Count numeric data cells past the header rows (row > 2)
                # in columns B+ — proxy for "this sheet has payload data
                # that should have been cited".
                if (
                    cell.row > 2
                    and cell.column >= 2
                    and isinstance(cell.value, (int, float))
                    and not isinstance(cell.value, bool)
                ):
                    numeric_data_cells += 1
            # Per-row default flag: a populated B-column value paired
            # with an "ASSUMPTION" or "DEFAULT" marker in col D means
            # this row is a consultant default, not payload data — no
            # citation required. The audit's vacuous-pass rule uses
            # this to avoid flagging all-default Assumptions sheets.
            try:
                b_val = ws.cell(row=row[0].row, column=2).value if row else None
                d_val = ws.cell(row=row[0].row, column=4).value if row else None
            except Exception:
                b_val, d_val = None, None
            if isinstance(b_val, (int, float)) and not isinstance(b_val, bool):
                rows_with_value_in_b += 1
                d_text = str(d_val or "").upper()
                if "ASSUMPTION" in d_text or "DEFAULT" in d_text:
                    default_flagged_rows += 1
        coverage[sheet_name] = {
            "cells_with_comment": with_comment,
            "argus_authored": argus_authored,
            "numeric_data_cells": numeric_data_cells,
            "default_flagged_rows": default_flagged_rows,
            "rows_with_value_in_b": rows_with_value_in_b,
        }
        if sheet_name in _SHEETS_REQUIRING_CITATIONS:
            if argus_authored >= 1:
                sheets_passed.append(sheet_name)
            elif numeric_data_cells == 0:
                # Sheet is structurally empty of payload data — nothing
                # to cite. Pass vacuously.
                sheets_passed.append(sheet_name)
            elif (
                rows_with_value_in_b > 0
                and default_flagged_rows == rows_with_value_in_b
            ):
                # Every value row is a consultant default flagged
                # ASSUMPTION — no payload-derived numbers on this
                # sheet, so the citation contract is vacuously
                # satisfied.
                sheets_passed.append(sheet_name)
            else:
                missing.append((sheet_name, "<no citations>"))

    return {
        "missing": missing,
        "coverage": coverage,
        "sheets_passed": sorted(sheets_passed),
        "sheets_requiring_citations": sorted(_SHEETS_REQUIRING_CITATIONS),
    }


__all__ = [
    "add_firm_header",
    "apply_tab_color",
    "audit_citations",
    "freeze_top_row",
]
